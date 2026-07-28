from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEETS: dict[str, list[tuple[str, str]]] = {
    "OCR_文本": [
        ("页码", "page_number"),
        ("总页数", "page_count"),
        ("文本块编号", "block_id"),
        ("类型", "block_type"),
        ("识别文本", "text"),
        ("置信度", "confidence"),
        ("坐标", "bbox"),
        ("证据编号", "evidence_id"),
    ],
    "OCR_表格": [
        ("页码", "page_number"),
        ("总页数", "page_count"),
        ("表格编号", "table_id"),
        ("行", "row"),
        ("列", "column"),
        ("跨行", "row_span"),
        ("跨列", "column_span"),
        ("识别文本", "text"),
        ("置信度", "confidence"),
        ("坐标", "bbox"),
        ("证据编号", "evidence_id"),
    ],
    "标准财务数据": [
        ("字段键", "field_key"),
        ("字段名称", "field_name"),
        ("期间", "period"),
        ("数值", "value"),
        ("单位", "unit"),
        ("证据编号", "evidence_id"),
    ],
    "识别问题": [
        ("问题类型", "issue_type"),
        ("问题说明", "message"),
        ("证据编号", "evidence_id"),
    ],
}

SOURCE_KEYS = {
    "OCR_文本": "text_blocks",
    "OCR_表格": "table_cells",
    "标准财务数据": "financial_data",
    "识别问题": "issues",
}


def _excel_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _style_sheet(sheet, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{max(sheet.max_row, 1)}"
    for column in range(1, column_count + 1):
        values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, sheet.max_row + 1)]
        width = min(max(max(map(len, values), default=8) + 2, 10), 48)
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _safe_sheet_title(page: int, table_id: str, used: set[str]) -> str:
    title = f"表_{page}_{table_id}".replace("/", "_").replace("\\", "_")[:31]
    candidate = title or "OCR表"
    suffix = 1
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{title[:31-len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _export_matrix_sheets(workbook: Workbook, normalized: dict[str, list[dict[str, Any]]]) -> None:
    """Export each OCR table as a real row/column matrix sheet.

    ``OCR_表格`` remains the evidence-level, one-cell-per-row contract.  The
    additional sheets are for human review and downstream table mapping: the
    OCR row/column coordinates are restored to their original matrix shape.
    """
    grouped: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for cell in normalized.get("table_cells", []):
        key = (int(cell.get("page_number", 0)), str(cell.get("table_id", "")))
        grouped.setdefault(key, []).append(cell)
    used = set(workbook.sheetnames)
    index = workbook.create_sheet("OCR_表格索引")
    index.append(["页码", "表格编号", "矩阵工作表", "行数", "列数"])
    for (page, table_id), cells in sorted(grouped.items()):
        title = _safe_sheet_title(page, table_id, used)
        sheet = workbook.create_sheet(title)
        max_row = max((int(cell.get("row", 0)) for cell in cells), default=0)
        max_column = max((int(cell.get("column", 0)) for cell in cells), default=0)
        matrix = [[None for _ in range(max_column)] for _ in range(max_row)]
        for cell in cells:
            row = int(cell.get("row", 0))
            column = int(cell.get("column", 0))
            if row > 0 and column > 0:
                matrix[row - 1][column - 1] = cell.get("text")
        for row in matrix:
            sheet.append(row)
        _style_sheet(sheet, max_column or 1)
        thin_gray = Side(style="thin", color="D9E2F3")
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
            for cell in row:
                cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        index.append([page, table_id, title, max_row, max_column])
    _style_sheet(index, 5)


def export_ocr_workbook(path: Path, normalized: dict[str, list[dict[str, Any]]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, columns in SHEETS.items():
        sheet = workbook.create_sheet(title)
        sheet.append([header for header, _ in columns])
        records = normalized.get(SOURCE_KEYS[title], [])
        for record in records:
            if isinstance(record, str):
                record = {"issue_type": "ocr", "message": record, "evidence_id": ""}
            sheet.append([_excel_value(record.get(key)) for _, key in columns])
        _style_sheet(sheet, len(columns))
    _export_matrix_sheets(workbook, normalized)
    workbook.save(path)
    return path


def read_ocr_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for title in SHEETS:
        sheet = workbook[title]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        result[title] = [
            dict(zip(headers, values, strict=True))
            for values in rows
            if any(value is not None for value in values)
        ]
    return result


def _json_value(value: Any) -> Any:
    """Restore list/dict values that were serialized into Excel cells."""
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text[0] not in "[{":
        return value
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return value


def normalized_from_ocr_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Load the stable OCR workbook contract back into pipeline records.

    This is deliberately separate from ``read_ocr_workbook`` so callers can
    reuse an earlier OCR result without invoking PaddleOCR again.
    """
    workbook = read_ocr_workbook(path)
    text_blocks = []
    for row in workbook.get("OCR_文本", []):
        text_blocks.append(
            {
                "page_number": row.get("页码"),
                "page_count": row.get("总页数"),
                "block_id": row.get("文本块编号"),
                "block_type": row.get("类型"),
                "text": row.get("识别文本"),
                "confidence": row.get("置信度"),
                "bbox": _json_value(row.get("坐标")),
                "evidence_id": row.get("证据编号"),
            }
        )
    table_cells = []
    for row in workbook.get("OCR_表格", []):
        table_cells.append(
            {
                "page_number": row.get("页码"),
                "page_count": row.get("总页数"),
                "table_id": row.get("表格编号"),
                "row": row.get("行"),
                "column": row.get("列"),
                "row_span": row.get("跨行") or 1,
                "column_span": row.get("跨列") or 1,
                "text": row.get("识别文本"),
                "confidence": row.get("置信度"),
                "bbox": _json_value(row.get("坐标")),
                "evidence_id": row.get("证据编号"),
            }
        )
    financial_data = []
    for row in workbook.get("标准财务数据", []):
        financial_data.append(
            {
                "field_key": row.get("字段键"),
                "field_name": row.get("字段名称"),
                "period": row.get("期间"),
                "value": _json_value(row.get("数值")),
                "unit": row.get("单位"),
                "evidence_id": row.get("证据编号"),
            }
        )
    issues = []
    for row in workbook.get("识别问题", []):
        issues.append(
            {
                "issue_type": row.get("问题类型"),
                "message": row.get("问题说明"),
                "evidence_id": row.get("证据编号"),
            }
        )
    return {
        "text_blocks": text_blocks,
        "table_cells": table_cells,
        "financial_data": financial_data,
        "issues": issues,
    }
