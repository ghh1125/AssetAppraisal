from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from demo.domain.financial_table_semantics import (
    CanonicalPeriod,
    canonical_long_term_asset_category,
    canonical_period,
    choose_historical_columns,
    detail_header_role,
)


def _text(value: Any) -> str:
    return re.sub(r"[\s：:()（）一二三四五六七八九十、．.]+", "", str(value or ""))


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("，", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _unit_scale_to_wan(sheet) -> float:
    samples: list[str] = [sheet.title]
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True):
        samples.extend(str(value) for value in row if value not in (None, ""))
    joined = " ".join(samples).replace(" ", "")
    if "万元" in joined:
        return 1.0
    if "金额单位" in joined and "元" in joined:
        return 0.0001
    if "（元）" in sheet.title or "(元)" in sheet.title:
        return 0.0001
    return 1.0


def _header_columns(sheet, row_number: int) -> tuple[int | None, int | None]:
    def scores(value: Any) -> tuple[int, int]:
        label = _text(str(value or "").replace("帐", "账"))
        if not label or any(token in label for token in ("增值", "减值", "增减")):
            return 0, 0
        book_score = 0
        appraised_score = 0
        if "账面" in label and "调整" not in label:
            if "净值" in label or "净额" in label:
                book_score = 30
            elif "原值" in label:
                book_score = 5
            elif "价值" in label or label.endswith("值"):
                book_score = 10
        if "评估" in label:
            if "净值" in label or "净额" in label:
                appraised_score = 30
            elif "原值" in label:
                appraised_score = 5
            elif "价值" in label or label.endswith("值"):
                appraised_score = 10
        return book_score, appraised_score

    best: tuple[int, int | None, int | None] = (0, None, None)
    for header_row in range(max(1, row_number - 100), row_number):
        variants: list[list[str]] = [
            [str(sheet.cell(header_row, column).value or "") for column in range(1, sheet.max_column + 1)]
        ]
        if header_row + 1 < row_number:
            inherited = ""
            combined: list[str] = []
            for column in range(1, sheet.max_column + 1):
                top = str(sheet.cell(header_row, column).value or "").strip()
                sub = str(sheet.cell(header_row + 1, column).value or "").strip()
                if top:
                    inherited = top
                combined.append(f"{top or inherited}{sub}" if sub else top)
            variants.append(combined)
        for labels in variants:
            book_candidates: list[tuple[int, int]] = []
            appraised_candidates: list[tuple[int, int]] = []
            for column, label in enumerate(labels, start=1):
                book_score, appraised_score = scores(label)
                if book_score:
                    book_candidates.append((book_score, column))
                if appraised_score:
                    appraised_candidates.append((appraised_score, column))
            if not book_candidates or not appraised_candidates:
                continue
            book_score, book_column = max(book_candidates)
            appraised_score, appraised_column = max(appraised_candidates)
            candidate = (
                book_score + appraised_score,
                book_column,
                appraised_column,
            )
            if candidate[0] > best[0]:
                best = candidate
    return best[1], best[2]


def _net_asset_candidate(sheet) -> dict[str, Any] | None:
    best: dict[str, Any] | None = None
    for row in sheet.iter_rows():
        for label_cell in row:
            label = _text(label_cell.value)
            if not label or not (
                label == "净资产"
                or "净资产所有者权益" in label
                or "所有者权益净资产" in label
            ):
                continue
            book_column, appraised_column = _header_columns(sheet, label_cell.row)
            if not book_column or not appraised_column:
                continue
            book = _number(sheet.cell(label_cell.row, book_column).value)
            appraised = _number(sheet.cell(label_cell.row, appraised_column).value)
            if book is None and appraised is None:
                continue
            score = 10
            if "汇总" in sheet.title:
                score += 5
            if "分类" not in sheet.title:
                score += 2
            candidate = {
                "score": score,
                "book": book,
                "appraised": appraised,
                "book_cell": f"{sheet.title}!{get_column_letter(book_column)}{label_cell.row}",
                "appraised_cell": f"{sheet.title}!{get_column_letter(appraised_column)}{label_cell.row}",
                "scale": _unit_scale_to_wan(sheet),
            }
            if best is None or candidate["score"] > best["score"]:
                best = candidate
    return best


def _equity_value_candidate(workbook) -> dict[str, Any] | None:
    best: dict[str, Any] | None = None
    for sheet in workbook.worksheets:
        scale = _unit_scale_to_wan(sheet)
        for row in sheet.iter_rows():
            for label_cell in row:
                label = _text(label_cell.value)
                if label != "股东全部权益价值":
                    continue
                numeric = None
                numeric_cell = None
                for column in range(label_cell.column + 1, min(sheet.max_column, label_cell.column + 8) + 1):
                    value = _number(sheet.cell(label_cell.row, column).value)
                    if value is not None:
                        numeric = value
                        numeric_cell = f"{sheet.title}!{get_column_letter(column)}{label_cell.row}"
                        break
                if numeric is None:
                    continue
                score = 20 + (5 if "净现金流" in sheet.title else 0)
                candidate = {
                    "score": score,
                    "value": numeric * scale,
                    "locator": numeric_cell,
                }
                if best is None or candidate["score"] > best["score"]:
                    best = candidate
    return best


def _products_and_tax(workbook) -> tuple[list[str], list[float], list[str]]:
    products: list[str] = []
    tax_rates: list[float] = []
    locators: list[str] = []
    for sheet in workbook.worksheets:
        if "主要产品" not in sheet.title and "产品及服务" not in sheet.title:
            continue
        header_row = None
        product_column = None
        tax_column = None
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
            for cell in row:
                label = _text(cell.value)
                if "产品或服务名称" in label:
                    header_row = cell.row
                    product_column = cell.column
                if "增值税率" in label:
                    tax_column = cell.column
            if header_row and product_column:
                break
        if not header_row or not product_column:
            continue
        for row_number in range(header_row + 1, min(sheet.max_row, header_row + 80) + 1):
            product = sheet.cell(row_number, product_column).value
            if isinstance(product, str) and product.strip() and product.strip() not in products:
                products.append(product.strip())
                locators.append(f"{sheet.title}!{get_column_letter(product_column)}{row_number}")
            if tax_column:
                rate = _number(sheet.cell(row_number, tax_column).value)
                if rate is not None:
                    normalized = rate / 100 if rate > 1 else rate
                    if normalized not in tax_rates:
                        tax_rates.append(normalized)
    return products, tax_rates, locators


def _percent(value: float) -> str:
    number = value * 100
    return f"{number:g}%"


def _canonical_scope_label(value: Any) -> str | None:
    label = _text(value)
    if not label:
        return None
    if label.startswith("其中"):
        label = label.removeprefix("其中")
    if label.endswith("净额"):
        label = label.removesuffix("净额")
    rules = (
        ("非流动资产", ("非流动资产",)),
        ("流动资产", ("流动资产",)),
        ("固定资产", ("固定资产",)),
        ("无形资产", ("无形资产",)),
        ("使用权资产", ("使用权资产",)),
        ("长期待摊费用", ("长期待摊费用",)),
        ("递延所得税资产", ("递延所得税资产",)),
        ("其他非流动资产", ("其他非流动资产",)),
        ("资产总计", ("资产总计", "资产合计")),
        ("非流动负债", ("非流动负债", "长期负债")),
        ("流动负债", ("流动负债",)),
        ("负债合计", ("负债合计", "负债总计")),
    )
    for canonical, aliases in rules:
        if any(label == alias or label.endswith(alias + "合计") for alias in aliases):
            return canonical
    if label == "净资产" or "净资产所有者权益" in label or "所有者权益净资产" in label:
        return "所有者权益"
    return None


def _summary_book_values(workbook) -> tuple[dict[str, float], dict[str, str]]:
    best_values: dict[str, float] = {}
    best_locators: dict[str, str] = {}
    best_score = -1
    for sheet in workbook.worksheets:
        if "汇总" not in sheet.title:
            continue
        values: dict[str, float] = {}
        locators: dict[str, str] = {}
        scale_to_yuan = _unit_scale_to_wan(sheet) * 10_000
        for row in sheet.iter_rows():
            for label_cell in row:
                canonical = _canonical_scope_label(label_cell.value)
                if not canonical:
                    continue
                book_column, _ = _header_columns(sheet, label_cell.row)
                if not book_column:
                    continue
                amount = _number(sheet.cell(label_cell.row, book_column).value)
                if amount is None:
                    continue
                values[canonical] = amount * scale_to_yuan
                locators[canonical] = (
                    f"{sheet.title}!{get_column_letter(book_column)}{label_cell.row}"
                )
        score = len(values) * 10
        if sheet.title in {"汇总表", "1-汇总表", "结果汇总"}:
            score += 5
        if any(name in sheet.title for name in ("固定资产", "流动负债", "非流动负债")):
            score -= 20
        if score > best_score:
            best_values, best_locators, best_score = values, locators, score
    return best_values, best_locators


def _detail_locator(sheet_title: str, column: int, rows: list[int]) -> str:
    letter = get_column_letter(column)
    if rows and rows == list(range(rows[0], rows[-1] + 1)):
        return f"{sheet_title}!{letter}{rows[0]}:{letter}{rows[-1]}"
    return "；".join(f"{sheet_title}!{letter}{row}" for row in rows)


def _detail_header_columns(
    sheet,
    header_row: int,
) -> tuple[dict[str, list[int]], dict[int, str], int]:
    direct: dict[str, list[int]] = {}
    labels: dict[int, str] = {}
    for cell in sheet[header_row]:
        role = detail_header_role(cell.value)
        if role:
            direct.setdefault(role, []).append(cell.column)
            labels[cell.column] = str(cell.value or "")
    next_row_has_value_subheaders = (
        header_row < sheet.max_row
        and any(
            _text(sheet.cell(header_row + 1, column).value)
            in {"原值", "净值", "净额"}
            for column in range(1, sheet.max_column + 1)
        )
    )
    if (
        len(direct.get("book_net", [])) == 1
        and not next_row_has_value_subheaders
    ):
        return direct, labels, header_row + 1
    if header_row >= sheet.max_row:
        return direct, labels, header_row + 1

    combined: dict[str, list[int]] = {}
    combined_labels: dict[int, str] = {}
    inherited_top = ""
    for column in range(1, sheet.max_column + 1):
        top = str(sheet.cell(header_row, column).value or "").strip()
        sub = str(sheet.cell(header_row + 1, column).value or "").strip()
        if top:
            inherited_top = top
        label = f"{top or inherited_top}{sub}" if sub else top
        role = detail_header_role(label)
        if role:
            combined.setdefault(role, []).append(column)
            combined_labels[column] = label
    return combined or direct, combined_labels or labels, header_row + 2


def _preferred_book_net_column(
    columns: list[int],
    labels: dict[int, str],
) -> int | None:
    preferred = [
        column
        for column in columns
        if "审计前" not in labels.get(column, "")
        and "调整前" not in labels.get(column, "")
    ]
    candidates = preferred or columns
    return candidates[0] if len(candidates) == 1 else None


def _electronic_equipment_detail_value(
    workbook,
) -> tuple[float | None, str, str, str]:
    best: tuple[int, float, str, str] | None = None
    ambiguous = False
    for sheet in workbook.worksheets:
        if not any(token in sheet.title for token in ("固定资产", "设备", "资产明细")):
            continue
        scale_to_yuan = _unit_scale_to_wan(sheet) * 10_000
        whole_sheet_is_electronic = (
            canonical_long_term_asset_category(sheet.title) == "电子设备"
            or any(
                canonical_long_term_asset_category(
                    sheet.cell(row, 1).value
                )
                == "电子设备"
                for row in range(1, min(sheet.max_row, 6) + 1)
            )
        )
        for header_row in range(1, min(sheet.max_row, 30) + 1):
            roles, labels, first_data_row = _detail_header_columns(
                sheet,
                header_row,
            )
            if not whole_sheet_is_electronic and len(roles.get("category", [])) != 1:
                continue
            amount_column = _preferred_book_net_column(
                roles.get("book_net", []),
                labels,
            )
            if amount_column is None:
                if len(roles.get("book_net", [])) > 1:
                    ambiguous = True
                continue
            category_column = (
                (roles.get("category") or [None])[0]
                if not whole_sheet_is_electronic
                else None
            )
            id_column = (roles.get("asset_id") or [None])[0]
            name_column = (roles.get("asset_name") or [None])[0]
            if id_column is None and name_column is None:
                continue
            amount = 0.0
            rows: list[int] = []
            for row_number in range(first_data_row, sheet.max_row + 1):
                category = (
                    "电子设备"
                    if category_column is None
                    else sheet.cell(row_number, category_column).value
                )
                if (
                    canonical_long_term_asset_category(category)
                    != "电子设备"
                ):
                    continue
                identity_values = [
                    sheet.cell(row_number, column).value
                    for column in (id_column, name_column)
                    if column is not None
                ]
                if any("合计" in str(value or "") for value in [category, *identity_values]):
                    continue
                if identity_values and not any(
                    str(value or "").strip() for value in identity_values
                ):
                    continue
                value = _number(sheet.cell(row_number, amount_column).value)
                if value is None:
                    continue
                amount += value * scale_to_yuan
                rows.append(row_number)
            if not rows:
                continue
            locator = _detail_locator(sheet.title, amount_column, rows)
            candidate = (len(rows), amount, locator, f"{len(rows)}项")
            if best is None or candidate[0] > best[0]:
                best = candidate
    if best is None:
        return (
            None,
            "",
            "",
            "ambiguous_candidate" if ambiguous else "source_absent",
        )
    _, amount, locator, quantity = best
    return amount, locator, quantity, ""


def _electronic_equipment_value(
    workbook,
) -> tuple[float | None, str, str, str]:
    summary_best: tuple[int, float, str] | None = None
    for sheet in workbook.worksheets:
        if "汇总" not in sheet.title:
            continue
        scale_to_yuan = _unit_scale_to_wan(sheet) * 10_000
        for row in sheet.iter_rows():
            for label_cell in row:
                if (
                    canonical_long_term_asset_category(label_cell.value)
                    != "电子设备"
                ):
                    continue
                book_column, _ = _header_columns(sheet, label_cell.row)
                if not book_column:
                    continue
                value = _number(sheet.cell(label_cell.row, book_column).value)
                if value is not None:
                    label = _text(label_cell.value)
                    score = (
                        (10 if "净额" in label or label.endswith("净值") else 0)
                        + (5 if "固定资产汇总" in sheet.title else 0)
                    )
                    candidate = (
                        score,
                        value * scale_to_yuan,
                        f"{sheet.title}!{get_column_letter(book_column)}{label_cell.row}",
                    )
                    if summary_best is None or candidate[0] > summary_best[0]:
                        summary_best = candidate
    detail_value, detail_locator, quantity, detail_reason = (
        _electronic_equipment_detail_value(workbook)
    )
    if summary_best is not None:
        _, value, locator = summary_best
        return value, locator, quantity, ""
    return detail_value, detail_locator, quantity, detail_reason


def _amount(value: float | None) -> str:
    return "XXX" if value is None else f"{value:,.2f}"


def _asset_tables(
    workbook,
) -> tuple[dict[str, Any], dict[str, dict[str, str]], list[str]]:
    values, locators = _summary_book_values(workbook)
    if not values:
        return {}, {}, []
    scope_labels = (
        ("流动资产账面金额：", "流动资产"),
        ("非流动资产账面金额：", "非流动资产"),
        ("其中：固定资产账面金额：", "固定资产"),
        ("无形资产账面金额：", "无形资产"),
        ("使用权资产账面金额：", "使用权资产"),
        ("长期待摊费用账面金额：", "长期待摊费用"),
        ("递延所得税资产账面金额：", "递延所得税资产"),
        ("其他非流动资产账面金额：", "其他非流动资产"),
        ("资产合计账面金额：", "资产总计"),
        ("流动负债账面金额：", "流动负债"),
        ("非流动负债账面金额：", "非流动负债"),
        ("负债合计账面金额：", "负债合计"),
        ("所有者权益账面金额：", "所有者权益"),
    )
    scope_rows = [[label, _amount(values.get(key))] for label, key in scope_labels]
    electronic, electronic_locator, electronic_quantity, electronic_reason = (
        _electronic_equipment_value(workbook)
    )
    long_rows = [
        ["项目", "账面金额（元）", "数量", "现状、特点"],
        [
            "电子设备",
            _amount(electronic),
            electronic_quantity,
            "以评估明细表为准",
        ],
        ["无形资产", _amount(values.get("无形资产")), "", "以评估明细表为准"],
        ["长期待摊费用", _amount(values.get("长期待摊费用")), "", "以评估明细表为准"],
    ]
    phrases = []
    for label, key in (
        ("固定资产", "固定资产"),
        ("无形资产", "无形资产"),
        ("长期待摊费用", "长期待摊费用"),
    ):
        if key in values:
            phrases.append(f"{label}账面价值{values[key]:,.2f}元")
    if electronic is not None:
        phrases.append(f"其中电子设备账面价值{electronic:,.2f}元")
    fields: dict[str, Any] = {
        "asset_scope_summary_table": {
            "caption": "被评估单位资产负债账面记录情况",
            "rows": scope_rows,
        },
        "long_term_assets_table": {
            "caption": "被评估单位主要长期资产账面记录情况",
            "rows": long_rows,
        },
    }
    if phrases:
        fields["major_long_term_assets"] = (
            "（一）被评估单位主要长期资产的账面记录情况如下：截至评估基准日，"
            + "；".join(phrases)
            + "。"
        )
    source_locator = "；".join(locators.values())
    evidence = {
        "asset_scope_summary_table": {
            "kind": "semantic_excel",
            "file": "",
            "locator": source_locator,
        },
        "long_term_assets_table": {
            "kind": "semantic_excel",
            "file": "",
            "locator": "；".join(filter(None, [source_locator, electronic_locator])),
        },
        "major_long_term_assets": {
            "kind": "semantic_excel",
            "file": "",
            "locator": "；".join(filter(None, [source_locator, electronic_locator])),
        },
    }
    issues: list[str] = []
    if electronic is None:
        reason = electronic_reason or "source_absent"
        issues.append(
            "long_term_assets_table："
            f"[{reason}] 未找到可唯一确定的电子设备账面净值，已保留 XXX"
        )
    for label, key in (
        ("无形资产", "无形资产"),
        ("长期待摊费用", "长期待摊费用"),
    ):
        if values.get(key) is None:
            issues.append(
                "long_term_assets_table："
                f"[source_absent] 未找到{label}账面金额，已保留 XXX"
            )
    return fields, evidence, issues


def _historical_header_parts(sheet, label_row: int) -> dict[int, list[Any]]:
    top_rows = range(1, min(label_row, 21))
    nearby_rows = range(max(1, label_row - 8), label_row)
    row_numbers = sorted(set(top_rows) | set(nearby_rows))
    headers: dict[int, list[Any]] = {}
    for column in range(1, sheet.max_column + 1):
        parts: list[Any] = []
        for row in row_numbers:
            value = sheet.cell(row, column).value
            if value in (None, "") or isinstance(value, (int, float, bool)):
                continue
            parts.append(value)
        headers[column] = parts
    return headers


def _period_columns_for_label(sheet, label_cell) -> list[tuple[int, CanonicalPeriod]]:
    headers = _historical_header_parts(sheet, label_cell.row)
    candidates: list[int] = []
    started = False
    for column in range(label_cell.column + 1, sheet.max_column + 1):
        selected = choose_historical_columns(
            {column: headers.get(column, [])},
            candidate_columns=[column],
            limit=1,
        )
        if selected:
            candidates.append(column)
            started = True
        elif started:
            break
    chosen = choose_historical_columns(
        headers,
        candidate_columns=candidates,
        limit=3,
    )
    result: list[tuple[int, CanonicalPeriod]] = []
    for column in chosen:
        periods = [
            period
            for value in headers.get(column, [])
            if (period := canonical_period(value)) is not None
        ]
        if not periods:
            continue
        period = next((item for item in periods if item[0] is not None), periods[-1])
        result.append((column, period))
    return result


def _balance_label(value: Any) -> str | None:
    label = _text(value)
    if label in {"资产总计", "资产合计"}:
        return "总资产"
    if label in {"负债合计", "负债总计"}:
        return "负债"
    if (
        label in {"净资产", "所有者权益合计", "所有者权益股东权益合计"}
        or "净资产所有者权益" in label
    ):
        return "所有者权益"
    return None


def _income_label(value: Any) -> str | None:
    label = _text(value)
    if "营业收入增长" in label or "营业收入比例" in label:
        return None
    checks = (
        ("一、营业收入", ("营业收入", "营业总收入")),
        ("减：营业成本", ("营业成本", "营业总成本")),
        ("税金及附加", ("税金及附加",)),
        ("销售费用", ("销售费用",)),
        ("管理费用", ("管理费用",)),
        ("研发费用", ("研发费用",)),
        ("财务费用", ("财务费用",)),
        ("加：投资收益", ("投资收益",)),
        ("二、营业利润", ("营业利润",)),
        ("加：营业外收入", ("营业外收入",)),
        ("减：营业外支出", ("营业外支出",)),
        ("三、利润总额", ("利润总额",)),
        ("减：所得税费用", ("所得税费用",)),
    )
    for canonical, aliases in checks:
        if any(label.endswith(alias) for alias in aliases):
            return canonical
    if (
        label.endswith("净利润")
        and not any(word in label for word in ("归属于", "持续经营", "终止经营", "净利润率"))
    ):
        return "四、净利润"
    return None


def _historical_table(
    workbook,
    *,
    kind: str,
) -> tuple[dict[str, Any] | None, str, bool]:
    matcher = _balance_label if kind == "balance" else _income_label
    preferred = ("资产负债表", "历资表") if kind == "balance" else ("利润表", "历利表")
    best: tuple[
        int,
        dict[str, dict[CanonicalPeriod, float]],
        list[CanonicalPeriod],
        list[str],
        bool,
    ] | None = None
    for sheet in workbook.worksheets:
        if not any(name in sheet.title for name in preferred):
            continue
        scale_to_yuan = _unit_scale_to_wan(sheet) * 10_000
        found: dict[str, dict[CanonicalPeriod, float]] = {}
        value_locators: dict[str, dict[CanonicalPeriod, str]] = {}
        period_order: list[CanonicalPeriod] = []
        locators: list[str] = []
        for row in sheet.iter_rows():
            for cell in row:
                canonical = matcher(cell.value)
                if not canonical:
                    continue
                values: dict[CanonicalPeriod, float] = {}
                cells: dict[CanonicalPeriod, str] = {}
                for column, period in _period_columns_for_label(sheet, cell):
                    amount = _number(sheet.cell(cell.row, column).value)
                    if amount is None:
                        continue
                    values[period] = amount * scale_to_yuan
                    cells[period] = (
                        f"{sheet.title}!{get_column_letter(column)}{cell.row}"
                    )
                    if period not in period_order:
                        period_order.append(period)
                if values and (
                    canonical not in found or len(values) > len(found[canonical])
                ):
                    found[canonical] = values
                    value_locators[canonical] = cells
                    locators.append(f"{sheet.title}!{cell.coordinate}")

        derived = False
        if kind == "balance" and "总资产" in found and "负债" in found:
            equity = dict(found.get("所有者权益", {}))
            equity_locators = dict(value_locators.get("所有者权益", {}))
            for period in set(found["总资产"]) & set(found["负债"]):
                if period in equity:
                    continue
                equity[period] = found["总资产"][period] - found["负债"][period]
                asset_cell = value_locators["总资产"][period]
                liability_cell = value_locators["负债"][period]
                equity_locators[period] = f"{asset_cell}－{liability_cell}"
                locators.extend([asset_cell, liability_cell])
                derived = True
            if equity:
                found["所有者权益"] = equity
                value_locators["所有者权益"] = equity_locators

        score = sum(len(values) for values in found.values())
        if sheet.title in preferred:
            score += 10
        if found and (best is None or score > best[0]):
            for cells in value_locators.values():
                locators.extend(cells.values())
            best = (
                score,
                found,
                period_order,
                list(dict.fromkeys(locators)),
                derived,
            )
    if best is None:
        return None, "", False
    _, found, period_order, locators, derived = best
    dated = sorted(
        (period for period in period_order if period[0] is not None),
        key=lambda period: (period[0] or 0, period[1] or 0, period[2] or 0),
    )
    relative = [period for period in period_order if period[0] is None]
    periods = ([*dated, *relative] if dated else relative)[-3:]
    defaults = ["历史期1", "历史期2", "评估基准期"]
    headers = [period[3] for period in periods]
    headers = [*defaults[: 3 - len(headers)], *headers]

    def formatted_row(label: str, values: dict[CanonicalPeriod, float] | None) -> list[str]:
        available = (
            []
            if values is None
            else [
                "XXX" if period not in values else f"{values[period]:,.2f}"
                for period in periods
            ]
        )
        return [label, *(["XXX"] * (3 - len(available))), *available]

    if kind == "balance":
        order = ("总资产", "负债", "所有者权益")
        title = "项目\\报表日"
        caption = "被评估单位近年资产负债状况见下表："
    else:
        order = (
            "一、营业收入",
            "减：营业成本",
            "税金及附加",
            "销售费用",
            "管理费用",
            "研发费用",
            "财务费用",
            "加：投资收益",
            "二、营业利润",
            "加：营业外收入",
            "减：营业外支出",
            "三、利润总额",
            "减：所得税费用",
            "四、净利润",
        )
        title = "项目\\报表年度"
        caption = "被评估单位近年经营状况见下表（利润表）："
    rows = [[title, *headers]]
    rows.extend(
        formatted_row(label, found.get(label))
        for label in order
    )
    return {"caption": caption, "rows": rows}, "；".join(locators), derived


def _workbook_valuation_method(workbook) -> str:
    market_tokens = ("市场法", "近期融资", "可比公司", "市盈率", "市净率")
    income_tokens = ("收益法", "现金流", "折现率", "永续期", "wacc")
    market_score = 0
    income_score = 0
    for sheet in workbook.worksheets:
        texts = [sheet.title]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, 120),
            min_col=1,
            max_col=min(sheet.max_column, 30),
            values_only=True,
        ):
            texts.extend(str(value) for value in row if isinstance(value, str))
        joined = "\n".join(texts).lower()
        market_score += sum(joined.count(token.lower()) for token in market_tokens)
        income_score += sum(joined.count(token.lower()) for token in income_tokens)
    if market_score > income_score:
        return "market"
    if income_score > market_score:
        return "income"
    return ""


def extract_workbook_facts(path: Path, role: str) -> dict[str, Any]:
    """Extract stable business facts without depending on a file name or cell address."""
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=True,
        keep_vba=path.suffix.lower() == ".xlsm",
    )
    fields: dict[str, Any] = {}
    evidence: dict[str, dict[str, str]] = {}
    issues: list[str] = []
    try:
        net_asset = None
        for sheet in workbook.worksheets:
            candidate = _net_asset_candidate(sheet)
            if candidate and (net_asset is None or candidate["score"] > net_asset["score"]):
                net_asset = candidate
        if role in {"reporting_workbook", "audited_financials"} and net_asset:
            if net_asset["book"] is not None:
                fields["book_net_assets"] = net_asset["book"] * net_asset["scale"]
                evidence["book_net_assets"] = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": net_asset["book_cell"],
                }
            if net_asset["appraised"] is not None:
                fields["asset_approach_value"] = net_asset["appraised"] * net_asset["scale"]
                evidence["asset_approach_value"] = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": net_asset["appraised_cell"],
                }
        if role in {"reporting_workbook", "audited_financials"}:
            table_fields, table_evidence, table_issues = _asset_tables(workbook)
            fields.update(table_fields)
            issues.extend(table_issues)
            for field_key, source in table_evidence.items():
                evidence[field_key] = {**source, "file": path.name}
            for history_kind, field_key in (
                ("balance", "historical_balance_sheet_table"),
                ("income", "historical_income_statement_table"),
            ):
                history, locator, derived = _historical_table(
                    workbook,
                    kind=history_kind,
                )
                if history:
                    fields[field_key] = history
                    evidence[field_key] = {
                        "kind": (
                            "semantic_excel_derived"
                            if derived
                            else "semantic_excel"
                        ),
                        "file": path.name,
                        "locator": locator,
                    }

        if role == "income_workbook":
            equity = _equity_value_candidate(workbook)
            if equity:
                fields["income_approach_value"] = equity["value"]
                evidence["income_approach_value"] = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": equity["locator"],
                }
            elif net_asset and net_asset["appraised"] is not None:
                value = net_asset["appraised"] * net_asset["scale"]
                source = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": net_asset["appraised_cell"],
                }
                method = _workbook_valuation_method(workbook)
                if method == "market":
                    fields["market_approach_value"] = value
                    evidence["market_approach_value"] = source
                elif method == "income":
                    fields["income_approach_value"] = value
                    evidence["income_approach_value"] = source

            products, tax_rates, product_locators = _products_and_tax(workbook)
            if products:
                fields["main_products"] = f"主要产品及服务：{'、'.join(products)}。"
                evidence["main_products"] = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": "；".join(product_locators),
                }
            if tax_rates:
                rates = "、".join(_percent(value) for value in tax_rates)
                fields["tax_rates"] = (
                    "被评估单位执行《企业会计准则》，主要产品及服务表列示的"
                    f"增值税税率为{rates}；其他税种及税率需结合税务资料复核。"
                )
                evidence["tax_rates"] = {
                    "kind": "semantic_excel",
                    "file": path.name,
                    "locator": "主要产品及服务!增值税率",
                }
            for history_kind, field_key in (
                ("balance", "historical_balance_sheet_table"),
                ("income", "historical_income_statement_table"),
            ):
                history, locator, derived = _historical_table(
                    workbook,
                    kind=history_kind,
                )
                if history:
                    fields[field_key] = history
                    evidence[field_key] = {
                        "kind": (
                            "semantic_excel_derived"
                            if derived
                            else "semantic_excel"
                        ),
                        "file": path.name,
                        "locator": locator,
                    }
    finally:
        workbook.close()
    return {"fields": fields, "evidence": evidence, "issues": issues}
