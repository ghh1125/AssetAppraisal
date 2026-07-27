from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from demo.domain.generation_issues import organize_generation_issues


DETAIL_COLUMNS = [
    ("status", "处理状态"),
    ("priority", "优先级"),
    ("page_number", "Word页码"),
    ("review_location", "检查位置"),
    ("field_name", "字段名称"),
    ("current_text", "当前内容"),
    ("problem", "问题说明"),
    ("expected_source", "应取来源"),
    ("source_file", "来源文件"),
    ("source_locator", "来源位置"),
    ("review_action", "处理建议"),
    ("category", "问题类别"),
    ("location_type", "位置类型"),
    ("location_id", "位置编号"),
    ("field_key", "标准字段"),
    ("page_basis", "页码依据"),
    ("issue_id", "问题编号"),
]


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    sheet.row_dimensions[1].height = 32


def _build_overview(workbook: Workbook, issues: list[dict]) -> None:
    sheet = workbook.active
    sheet.title = "检查总览"
    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = "资产评估报告生成问题检查总览"
    title.font = Font(size=16, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34

    priority_counts = Counter(str(item.get("priority", "")) for item in issues)
    pages = [
        str(item.get("page_number"))
        for item in issues
        if item.get("page_number") not in (None, "")
    ]
    metrics = [
        ("待处理问题总数", len(issues)),
        ("高优先级", priority_counts.get("高", 0)),
        ("中优先级", priority_counts.get("中", 0)),
        ("低优先级", priority_counts.get("低", 0)),
        ("涉及 Word 页数", len(set(pages))),
    ]
    for row_number, (label, value) in enumerate(metrics, start=3):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, value)
        sheet.cell(row_number, 1).font = Font(bold=True)
        sheet.cell(row_number, 1).fill = PatternFill(
            "solid",
            fgColor="D9EAF7",
        )
        sheet.cell(row_number, 2).alignment = Alignment(
            horizontal="center"
        )

    sheet.merge_cells("D3:F3")
    sheet["D3"] = "检查步骤"
    sheet["D3"].font = Font(bold=True, color="FFFFFF")
    sheet["D3"].fill = PatternFill("solid", fgColor="70AD47")
    instructions = [
        "1. 按“问题明细”的 Word 页码升序打开报告。",
        "2. 根据“检查位置”和“当前内容”找到黄色占位符。",
        "3. 按“应取来源、来源文件、来源位置、处理建议”补充或确认。",
    ]
    for row_number, instruction in enumerate(instructions, start=4):
        sheet.merge_cells(
            start_row=row_number,
            start_column=4,
            end_row=row_number,
            end_column=6,
        )
        sheet.cell(row_number, 4, instruction)
        sheet.cell(row_number, 4).alignment = Alignment(wrap_text=True)

    page_counts = Counter(
        str(item.get("page_number") or "页码待确认")
        for item in issues
    )
    sheet["A10"] = "按页码汇总"
    sheet["A10"].font = Font(bold=True, color="FFFFFF")
    sheet["A10"].fill = PatternFill("solid", fgColor="4472C4")
    sheet["A11"] = "Word页码"
    sheet["B11"] = "问题数量"
    for cell in sheet[11][:2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    ordered_pages = sorted(
        page_counts,
        key=lambda value: (
            value == "页码待确认",
            int(value) if value.isdigit() else 10**9,
        ),
    )
    for row_number, page in enumerate(ordered_pages, start=12):
        sheet.cell(row_number, 1, page)
        sheet.cell(row_number, 2, page_counts[page])
        sheet.cell(row_number, 2).alignment = Alignment(
            horizontal="center"
        )

    for column, width in enumerate((24, 14, 4, 28, 28, 28), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False


def _build_detail(workbook: Workbook, issues: list[dict]) -> None:
    sheet = workbook.create_sheet("问题明细")
    sheet.append([label for _, label in DETAIL_COLUMNS])
    for issue in issues:
        sheet.append(
            [issue.get(key, "") for key, _ in DETAIL_COLUMNS]
        )
    _style_header(sheet)
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    widths = [
        14,
        9,
        10,
        44,
        24,
        18,
        38,
        20,
        30,
        38,
        44,
        22,
        14,
        28,
        28,
        18,
        18,
    ]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for column in range(12, 18):
        dimension = sheet.column_dimensions[get_column_letter(column)]
        dimension.hidden = True
        dimension.outlineLevel = 1

    thin_gray = Side(style="thin", color="D9E1F2")
    priority_fills = {
        "高": PatternFill("solid", fgColor="F4CCCC"),
        "中": PatternFill("solid", fgColor="FFF2CC"),
        "低": PatternFill("solid", fgColor="D9EAD3"),
    }
    pending_fill = PatternFill("solid", fgColor="FCE4D6")
    alternate_fill = PatternFill("solid", fgColor="F7F9FC")
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2),
        start=2,
    ):
        if row_number % 2 == 0:
            for cell in row:
                cell.fill = alternate_fill
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = Border(bottom=thin_gray)
        priority = str(row[1].value or "")
        if priority in priority_fills:
            row[1].fill = priority_fills[priority]
            row[1].font = Font(bold=True)
        if str(row[0].value or "") == "待人工处理":
            row[0].fill = pending_fill
            row[0].font = Font(bold=True)
        row[2].alignment = Alignment(
            horizontal="center",
            vertical="top",
        )
        sheet.row_dimensions[row_number].height = 42


def export_generation_issues(
    output: Path,
    issues: list[dict],
) -> Path:
    organized = organize_generation_issues(issues)
    workbook = Workbook()
    _build_overview(workbook, organized)
    _build_detail(workbook, organized)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
