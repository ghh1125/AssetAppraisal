from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_audit(
    output: Path,
    locations: list[dict],
    fields: dict[str, object],
    sources: dict[str, dict],
    template_pages: dict[str, int | str] | None = None,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "填充结果"
    headers = ["位置编号", "原模板页码", "类型", "原文上下文", "标准字段", "字段名称", "最终填充值", "来源类别", "来源文件", "来源位置"]
    sheet.append(headers)
    for item in locations:
        source = sources.get(item["field_key"], {})
        sheet.append([
            item["location_id"], (template_pages or {}).get(item["location_id"], ""), item["record_type"], item["context"], item["field_key"], item["field_name"],
            str(fields.get(item["field_key"], "")), source.get("kind", "fallback"), source.get("file", ""), source.get("locator", ""),
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [22, 12, 14, 45, 28, 24, 45, 18, 38, 45]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def write_json(path: Path, payload) -> Path:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path
