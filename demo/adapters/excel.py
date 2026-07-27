from __future__ import annotations

from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def read_cells(path: Path, locators: list[str]) -> dict[str, object]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        result = {}
        for locator in locators:
            sheet, coordinate = locator.rsplit("!", 1)
            result[locator] = workbook[sheet][coordinate].value
        return result
    finally:
        workbook.close()


def try_read_cells(
    path: Path | None,
    locators: list[str],
) -> tuple[dict[str, object], list[str]]:
    if path is None:
        return {}, ["缺少来源文件"]
    try:
        return read_cells(path, locators), []
    except KeyError as exc:
        return {}, [f"缺少工作表：{exc}"]
    except (OSError, ValueError, BadZipFile) as exc:
        return {}, [f"材料无法读取：{exc}"]


def read_range_values(path: Path, locator: str) -> list[object]:
    """Read a rectangular Excel range into a flat row-major list."""
    sheet_name, cell_range = locator.rsplit("!", 1)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name]
        return [
            sheet.cell(row=row, column=column).value
            for row in range(min_row, max_row + 1)
            for column in range(min_col, max_col + 1)
        ]
    finally:
        workbook.close()


def find_labels(path: Path, labels: list[str]) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    found = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    text = str(cell.value or "").strip()
                    if any(label in text for label in labels):
                        found.append({"sheet": sheet.title, "coordinate": cell.coordinate, "value": cell.value})
        return found
    finally:
        workbook.close()


def _format_table_value(value: object, blank: str = "") -> str:
    if value in (None, ""):
        return blank
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return str(value).strip()


def read_configured_table(path: Path, spec: dict) -> list[list[str]]:
    """Read one configured financial table as a display-ready string matrix."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[spec["sheet"]]
        matrix = [list(spec["header"])] if spec.get("include_header", True) else []
        for row in spec["rows"]:
            values = [
                _format_table_value(sheet[coordinate].value, row.get("blank", ""))
                for coordinate in row["cells"]
            ]
            matrix.append([row["label"], *values])
        return matrix
    finally:
        workbook.close()


def try_read_configured_table(
    path: Path | None,
    spec: dict,
) -> tuple[list[list[str]] | None, list[str]]:
    if path is None:
        return None, ["缺少来源文件"]
    try:
        return read_configured_table(path, spec), []
    except KeyError as exc:
        return None, [f"缺少工作表：{exc}"]
    except (OSError, ValueError, BadZipFile) as exc:
        return None, [f"材料无法读取：{exc}"]
