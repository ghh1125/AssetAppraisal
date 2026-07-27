from openpyxl import load_workbook

from demo.adapters.generation_issues import export_generation_issues
from demo.domain.generation_issues import (
    issues_for_missing_locations,
    issues_from_word_findings,
    organize_generation_issues,
)
from demo.domain.replacement import build_replacements


def test_missing_field_uses_original_marker_and_creates_location_issue():
    location = {
        "location_id": "DOCUMENT-P0001-X01",
        "record_type": "占位符",
        "field_key": "company_name",
        "field_name": "公司名称",
        "marker": "XXX",
        "context": "公司名称：XXX",
        "source_kind": "人工输入",
        "source_file": "人工基础信息",
        "source_locator": "公司名称",
    }

    replacements = build_replacements([location], {})
    issues = issues_for_missing_locations([location], {})

    assert replacements[location["location_id"]] == "XXX"
    assert issues[0]["location_id"] == location["location_id"]
    assert issues[0]["current_text"] == "XXX"
    assert issues[0]["expected_source"] == "人工输入"


def test_missing_20xx_field_preserves_full_marker():
    location = {
        "location_id": "DOCUMENT-P0001-X01",
        "field_key": "valuation_year",
        "field_name": "评估年份",
        "marker": "20XX",
        "context": "评估基准日：20XX年",
    }

    assert build_replacements([location], {})[location["location_id"]] == "20XX"


def test_generation_issue_export_has_required_columns(tmp_path):
    output = export_generation_issues(
        tmp_path / "生成问题清单.xlsx",
        [
            {
                "issue_id": "GEN-1",
                "priority": "高",
                "category": "missing_field",
                "page_number": 9,
                "page_basis": "generated_report",
                "location_id": "DOCUMENT-P0100-X01",
                "location_type": "段落",
                "location_description": "公司名称：XXX",
                "field_key": "company_name",
                "field_name": "公司名称",
                "current_text": "XXX",
                "problem": "指定来源未匹配到可用值",
                "expected_source": "人工输入",
                "source_file": "",
                "source_locator": "公司名称",
                "suggestion": "补充公司名称",
                "status": "待人工处理",
            }
        ],
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["检查总览", "问题明细"]
    overview = workbook["检查总览"]
    detail = workbook["问题明细"]
    assert overview["A1"].value == "资产评估报告生成问题检查总览"
    assert overview["B3"].value == 1
    assert "D3:F3" in {
        str(cell_range) for cell_range in overview.merged_cells.ranges
    }
    assert detail["C2"].value == 9
    assert detail["D2"].value == "第9页｜公司名称：XXX"
    assert detail["N2"].value == "DOCUMENT-P0100-X01"
    assert detail.freeze_panes == "E2"
    assert detail.column_dimensions["L"].hidden is True
    assert detail.column_dimensions["Q"].hidden is True


def test_generation_issues_are_sorted_for_page_by_page_review():
    organized = organize_generation_issues(
        [
            {
                "page_number": "",
                "priority": "高",
                "location_id": "C",
                "location_description": "未定位",
                "suggestion": "确认位置",
            },
            {
                "page_number": 9,
                "priority": "中",
                "location_id": "B",
                "location_description": "利润表",
                "suggestion": "补利润表",
            },
            {
                "page_number": 5,
                "priority": "高",
                "location_id": "A",
                "location_description": "公司名称",
                "suggestion": "补公司名称",
            },
        ]
    )

    assert [item["page_number"] for item in organized] == [5, 9, ""]
    assert organized[0]["review_location"] == "第5页｜公司名称"
    assert organized[-1]["review_location"] == "页码待确认｜未定位"
    assert organized[1]["review_action"] == "补利润表"


def test_word_finding_keeps_mapped_field_source():
    issues = issues_from_word_findings(
        [
            {
                "location_id": "DOCUMENT-P0005-X01",
                "part": "word/document.xml",
                "paragraph_index": 5,
                "occurrence_index": 1,
                "location_type": "段落",
                "context": "XXX有限责任公司拟收购",
                "current_text": "XXX",
            }
        ],
        [
            {
                "location_id": "DOCUMENT-P0005-X01",
                "field_key": "commissioning_party_name",
                "field_name": "委托方名称",
                "source_kind": "人工输入",
                "source_locator": "委托方名称",
            }
        ],
        {"commissioning_party_name": ""},
        {"commissioning_party_name": {"kind": "missing"}},
    )

    assert issues[0]["field_key"] == "commissioning_party_name"
    assert issues[0]["expected_source"] == "人工输入"


def test_unfinished_appraisal_has_specific_review_message():
    issues = issues_from_word_findings(
        [
            {
                "location_id": "DOCUMENT-P0089-X02",
                "part": "word/document.xml",
                "paragraph_index": 89,
                "occurrence_index": 2,
                "location_type": "段落",
                "context": "评估结论为XX万元",
                "current_text": "XX",
            }
        ],
        [
            {
                "location_id": "DOCUMENT-P0089-X02",
                "field_key": "asset_approach_value",
                "field_name": "资产基础法评估值",
                "source_kind": "Excel",
            }
        ],
        {"asset_approach_value": ""},
        {
            "asset_approach_value": {
                "kind": "unfinished_appraisal",
                "file": "资产清查.xlsx",
                "locator": "汇总表!D22",
            }
        },
    )

    assert issues[0]["category"] == "unfinished_appraisal"
    assert issues[0]["problem"] == "疑似尚未完成评估，评估列为空或全零"
    assert "保持黄色" in issues[0]["suggestion"]
    assert issues[0]["source_file"] == "资产清查.xlsx"
    assert issues[0]["source_locator"] == "汇总表!D22"
