from pathlib import Path

from openpyxl import Workbook

from demo.adapters.company_api import (
    CompanyApiAdapter,
    QichachaApiAdapter,
    comparable_search_terms,
)
from demo.adapters.bailian_glm import BailianYellowNarrativeAdapter
from demo.adapters.excel import read_cells, try_read_cells
from demo.adapters.llm import LlmAdapter
from demo.adapters.ocr import OcrAdapter


def test_excel_adapter_reads_exact_cells_without_modifying_source(tmp_path: Path):
    path = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "项目信息"
    ws["B5"] = "示例有限公司"
    wb.save(path)
    before = path.read_bytes()
    assert read_cells(path, ["项目信息!B5"])["项目信息!B5"] == "示例有限公司"
    assert path.read_bytes() == before


def test_optional_excel_cell_returns_issue_instead_of_raising(tmp_path: Path):
    workbook_path = tmp_path / "other-layout.xlsx"
    workbook = Workbook()
    workbook.active.title = "资产负债表"
    workbook.save(workbook_path)

    values, issues = try_read_cells(
        workbook_path,
        ["06N_资产负债表!L75"],
    )

    assert values == {}
    assert "缺少工作表" in issues[0]


def test_optional_adapters_degrade_without_clients_or_credentials(tmp_path: Path):
    assert OcrAdapter().extract(tmp_path / "missing.pdf")[0] == []
    assert CompanyApiAdapter().fetch("示例有限公司")[0] == {}
    assert LlmAdapter().generate({"name": "示例有限公司"})[0] == {}


class _Response:
    def __init__(self, payload):
        self.payload = payload

    def raise_for_status(self):
        return None

    def json(self):
        return self.payload


class _Client:
    def __init__(self):
        self.calls = []

    def get(self, url, **kwargs):
        self.calls.append((url, kwargs))
        if url.endswith("ECIInfoVerify/GetInfo"):
            return _Response({"Status": "200", "Data": {"Name": "示例有限公司", "Partners": [{"StockName": "甲", "StockPercent": "60%"}], "ChangeRecords": [{"ChangeDate": "2024-01-01", "ProjectName": "股东变更", "BeforeContent": "甲", "AfterContent": "乙"}]}})
        if url.endswith("tm/SearchByApplicant"):
            return _Response({"Status": "200", "Result": [{"Name": "示例商标", "RegNo": "123", "Category": "35", "FlowStatusDesc": "注册"}]})
        if url.endswith("PatentV4/Search") or url.endswith("PatentV4/SearchMultiPatents"):
            return _Response({"Status": "200", "Data": [{"Title": "示例专利", "PatentType": "发明", "LegalStatus": "有效"}]})
        return _Response({"Status": "200", "Result": [{"Name": "示例软件", "RegisterNo": "2024SR0001", "VersionNo": "V1.0"}]})


def test_qichacha_adapter_signs_and_maps_only_authorized_fields():
    client = _Client()
    adapter = QichachaApiAdapter(client, "app", "secret", extra_api_codes=())
    payload, issues = adapter.fetch("示例有限公司")
    assert not issues
    assert set(payload["fields"]) == {
        "commissioning_party_profile", "ownership_history", "ownership_at_valuation_date",
        "unrecorded_intangibles", "software_copyrights",
    }
    assert "统一社会信用代码" not in payload["fields"]["commissioning_party_profile"]
    assert "示例专利" in payload["fields"]["unrecorded_intangibles"]
    assert "示例商标" in payload["fields"]["unrecorded_intangibles"]
    assert "示例软件" in payload["fields"]["software_copyrights"]
    assert len(client.calls) == 4
    first_url, first_kwargs = client.calls[0]
    assert first_url.endswith("ECIInfoVerify/GetInfo")
    timespan = first_kwargs["headers"]["Timespan"]
    assert first_kwargs["headers"]["Token"] == QichachaApiAdapter.token("app", timespan, "secret")


def test_graphical_trademark_uses_text_name_and_marks_software_query_successfully():
    class GraphicClient(_Client):
        def get(self, url, **kwargs):
            if url.endswith("tm/SearchByApplicant"):
                return _Response({"Status": "200", "Result": [{"ImageUrl": "https://example.test/mark.png", "RegNo": "1"}]})
            if url.endswith("CopyRight/SearchCopyRight"):
                return _Response({"Status": "200", "Result": []})
            return super().get(url, **kwargs)

    payload, issues = QichachaApiAdapter(GraphicClient(), "app", "secret", extra_api_codes=()).fetch("示例有限公司")
    assert not issues
    assert payload["trademark_rows"][0]["name"] == "图形"
    assert payload["trademark_rows"][0]["image"].startswith("https://")
    assert payload["software_query_ok"] is True


def test_qichacha_non_review_apis_are_default_and_expose_evidence():
    class ExtendedClient(_Client):
        def get(self, url, **kwargs):
            if url.endswith("EnterpriseInfo/Verify"):
                self.calls.append((url, kwargs))
                return _Response({"Status": "200", "Data": {"Name": "示例有限公司", "QccIndustry": "电子元件", "Scope": "芯片封装测试"}})
            if url.endswith("AR/GetAnnualReport"):
                self.calls.append((url, kwargs))
                return _Response({"Status": "200", "Data": [{"Year": "2025", "Asset": "100"}]})
            return super().get(url, **kwargs)

    client = ExtendedClient()
    adapter = QichachaApiAdapter(client, "app", "secret")
    payload, issues = adapter.fetch("示例有限公司")
    assert not issues
    assert {item["api_code"] for item in payload["evidence"]} >= {"2001", "213"}
    assert len(client.calls) == 6


def test_qichacha_default_does_not_call_optional_paid_apis():
    client = _Client()
    payload, issues = QichachaApiAdapter(client, "app", "secret", extra_api_codes=()).fetch("示例有限公司")
    assert not issues
    assert "evidence" not in payload or not any(item["api_code"] in {"2001", "213"} for item in payload["evidence"])
    assert len(client.calls) == 4


def test_qichacha_default_calls_all_non_review_apis():
    client = _Client()
    payload, issues = QichachaApiAdapter(client, "app", "secret").fetch("示例有限公司")
    assert not issues
    called_paths = {url.split("api.qichacha.com/", 1)[-1] for url, _ in client.calls}
    assert called_paths == {
        "ECIInfoVerify/GetInfo", "tm/SearchByApplicant", "PatentV4/Search", "CopyRight/SearchCopyRight",
        "EnterpriseInfo/Verify", "AR/GetAnnualReport",
    }
    assert len(client.calls) == 6


def test_qichacha_does_not_register_review_only_apis():
    assert "962" not in QichachaApiAdapter.DEFAULT_ENDPOINTS


def test_qichacha_discovers_listed_comparable_candidates_from_announcements_without_unverified_detail_calls():
    class ComparableClient(_Client):
        def get(self, url, **kwargs):
            self.calls.append((url, kwargs))
            if url.endswith("FuzzySearch/GetList"):
                return _Response({"Status": "200", "Result": [{"Name": "同行企业甲", "Industry": "汽车零部件"}]})
            if url.endswith("IPOAnnouncement/GetList"):
                return _Response({"Status": "200", "Result": {"Data": [{"CompanyName": "上市公司甲", "KeyNo": "peer-key", "StockCode": "600001", "Title": "热处理业务公告", "PublishDate": "2026-01-01"}]}})
            return super().get(url, **kwargs)

    adapter = QichachaApiAdapter(ComparableClient(), "app", "secret")
    evidence, issues = adapter.discover_listed_comparables(["汽车零部件热处理"])

    assert issues == []
    assert {item["api_code"] for item in evidence} == {"886", "915"}
    assert "上市公司甲" in next(item["text"] for item in evidence if item["api_code"] == "915")
    assert "600001" in next(item["text"] for item in evidence if item["api_code"] == "915")
    assert all("IPO/GetIPO" not in url for url, _ in adapter.client.calls)


def test_listed_announcements_are_deduplicated_by_company_and_stock_code():
    from demo.adapters.company_api import _listed_announcements_text

    text = _listed_announcements_text(
        {"Status": "200", "Result": {"Data": [
            {"CompanyName": "上市公司甲", "StockCode": "600001", "Title": "公告一", "PublishDate": "2026-01-01"},
            {"CompanyName": "上市公司甲", "StockCode": "600001", "Title": "公告二", "PublishDate": "2026-01-02"},
            {"CompanyName": "上市公司乙", "StockCode": "600002", "Title": "公告三", "PublishDate": "2026-01-03"},
        ]}}
    )

    assert text.count("上市公司甲") == 1
    assert "上市公司乙" in text


def test_qichacha_can_disable_paid_comparable_discovery():
    client = _Client()
    adapter = QichachaApiAdapter(
        client,
        "app",
        "secret",
        enable_comparable_discovery=False,
    )
    evidence, issues = adapter.discover_listed_comparables(["汽车零部件热处理"])
    assert evidence == []
    assert issues == []
    assert client.calls == []


def test_comparable_search_terms_reduce_business_scope_to_compact_industry_phrases():
    assert comparable_search_terms(
        "各类汽车零部件的热处理加工并提供相关的技术开发、技术咨询及售后服务。"
    ) == ["汽车零部件热处理", "汽车零部件", "热处理"]


def test_comparable_narrative_evidence_requires_an_announcement_with_multiple_dimensions():
    evidence = [
        {
            "evidence_id": "api:qichacha:target:735:profile",
            "text": "被评估单位工商信息：行业：制造业；经营范围：汽车零部件热处理",
        },
        {
            "evidence_id": "api:qichacha:target:915:peer:汽车零部件热处理",
            "text": "上市公司公告候选：上市公司甲；股票代码：600001；公告：热处理业务公告；日期：2026-01-01",
        },
    ]

    selected = BailianYellowNarrativeAdapter._relevant_evidence(
        "comparable_list", evidence
    )

    assert selected == [evidence[1]]


def test_comparable_narrative_evidence_rejects_a_name_only_candidate():
    evidence = [
        {
            "evidence_id": "api:qichacha:target:915:peer:工业设备",
            "text": "上市公司公告候选：上市公司甲；股票代码：600001",
        },
    ]

    selected = BailianYellowNarrativeAdapter._relevant_evidence(
        "comparable_list", evidence
    )

    assert selected == []


def test_narrative_prompt_forbids_model_world_knowledge():
    prompt = Path("demo/prompts/yellow_narratives.v2.txt").read_text(encoding="utf-8")
    assert "不得使用模型自身知识" in prompt
    assert "未披露主要客户及供应商" in prompt
    assert "不等于可比性最终认定" in prompt
