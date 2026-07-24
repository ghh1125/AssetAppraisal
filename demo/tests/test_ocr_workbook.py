from copy import deepcopy

from openpyxl import load_workbook

from demo.adapters.ocr_workbook import export_ocr_workbook, normalized_from_ocr_workbook, read_ocr_workbook


def test_exports_four_auditable_sheets_without_mutating_input(tmp_path):
    normalized = {
        "text_blocks": [
            {
                "page_number": 1,
                "page_count": 2,
                "block_id": "p1-b1",
                "block_type": "text",
                "text": "利润表",
                "confidence": 0.9,
                "bbox": [1, 2, 3, 4],
                "evidence_id": "pdf:p1:b1",
            }
        ],
        "table_cells": [
            {
                "page_number": 2,
                "page_count": 2,
                "table_id": "p2-t1",
                "row": 3,
                "column": 4,
                "text": "1,234.50",
                "confidence": 0.95,
                "bbox": [5, 6, 7, 8],
                "evidence_id": "pdf:p2:t1:r3:c4",
            }
        ],
        "financial_data": [
            {
                "field_key": "total_assets",
                "field_name": "资产总计",
                "period": "2025-06-30",
                "value": 1234.5,
                "unit": "元",
                "evidence_id": "pdf:p2:t1:r3:c4",
            }
        ],
        "issues": [{"issue_type": "low_confidence", "message": "请复核", "evidence_id": "pdf:p1:b1"}],
    }
    before = deepcopy(normalized)

    path = export_ocr_workbook(tmp_path / "OCR结构化结果.xlsx", normalized)
    reloaded = read_ocr_workbook(path)

    assert normalized == before
    assert set(reloaded) == {"OCR_文本", "OCR_表格", "标准财务数据", "识别问题"}
    assert reloaded["OCR_文本"][0]["证据编号"] == "pdf:p1:b1"
    assert reloaded["OCR_表格"][0]["页码"] == 2
    assert reloaded["OCR_表格"][0]["行"] == 3
    assert reloaded["标准财务数据"][0]["数值"] == 1234.5

    workbook = load_workbook(path)
    assert "OCR_表格索引" in workbook.sheetnames
    matrix_sheet = next(name for name in workbook.sheetnames if name.startswith("表_2_p2-t1"))
    assert workbook[matrix_sheet].cell(row=3, column=4).value == "1,234.50"
    for sheet in workbook.worksheets:
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref
        assert sheet.row_dimensions[1].height >= 24


def test_normalized_ocr_workbook_round_trip_restores_pipeline_contract(tmp_path):
    path = export_ocr_workbook(
        tmp_path / "OCR结构化结果.xlsx",
        {
            "text_blocks": [{"page_number": 1, "page_count": 1, "block_id": "p1-b1", "block_type": "text", "text": "审计报告", "confidence": 0.9, "bbox": [1, 2, 3, 4], "evidence_id": "pdf:p1:b1"}],
            "table_cells": [],
            "financial_data": [],
            "issues": [],
        },
    )
    normalized = normalized_from_ocr_workbook(path)
    assert normalized["text_blocks"][0]["text"] == "审计报告"
    assert normalized["text_blocks"][0]["bbox"] == [1, 2, 3, 4]
    assert normalized["text_blocks"][0]["evidence_id"] == "pdf:p1:b1"
