from __future__ import annotations

import hashlib
import json
import re
import shutil
import tempfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from demo import schemas
from demo.adapters.audit import export_audit, write_json
from demo.adapters.document import read_narrative_evidence, read_table_matrix
from demo.adapters.excel import (
    read_cells,
    try_read_cells,
    try_read_configured_table,
)
from demo.adapters.generation_issues import export_generation_issues
from demo.adapters.ocr_workbook import export_ocr_workbook, normalized_from_ocr_workbook
from demo.adapters.review_inputs import build_data_review_evidence, build_format_review_evidence, build_semantic_review_evidence
from demo.adapters.workflow_trace import (
    WorkflowTraceRecorder,
    trace_candidates,
    trace_evidence,
    trace_mappings,
    trace_resolved_fields,
)
from demo.adapters.word import (
    document_paragraph_texts,
    fill_template,
    highlight_unresolved_placeholders,
    inventory_template,
    replace_image_markers,
    replace_report_number_year,
)
from demo.domain.generation_issues import (
    apply_page_locations,
    issues_from_word_findings,
)
from demo.domain.review import aggregate_reviews
from demo.domain.field_validation import (
    apply_missing_field_policy,
    normalize_narrative_modules,
    normalize_valuation_methods,
    require_financial_fields,
)
from demo.domain.field_validation import validate_valuation_subject_type
from demo.domain.financial_matching import blank_configured_table
from demo.domain.mapping import validate_mapping
from demo.domain.narrative_policy import (
    select_narrative_fields,
    should_create_candidate_report,
)
from demo.domain.ocr_normalization import normalize_ocr_pages
from demo.domain.pdf_ocr_fields import find_ocr_table, resolve_configured_ocr_fields, resolve_ocr_aux_fields
from demo.domain.replacement import build_replacements
from demo.domain.template_pagination import map_location_pages
from demo.domain.workflow_contracts import validate_workflow_contract
from demo.domain.yellow_routing import (
    RouteKind,
    fields_for_route,
    load_yellow_routes,
    validate_yellow_routes,
)
from demo.run import run_project


@dataclass(frozen=True)
class PipelineResult:
    report_path: Path
    audit_path: Path
    ocr_workbook_path: Path | None
    manifest_path: Path
    issues: list[str]


OcrFieldResolver = Callable[
    [dict[str, list[dict[str, Any]]], dict[str, Any]],
    tuple[dict[str, Any], list[str]],
]


def _path(base: Path, value: str | Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else (base / path).resolve()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _legacy_evidence(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["填充结果"]
    header = {str(cell.value): index for index, cell in enumerate(sheet[1])}
    result: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        field_key = str(row[header["标准字段"]])
        result.setdefault(
            field_key,
            {
                "kind": str(row[header["来源类别"]] or "legacy"),
                "file": str(row[header["来源文件"]] or ""),
                "locator": str(row[header["来源位置"]] or ""),
            },
        )
    return result


def _default_ocr_field_resolver(
    normalized: dict[str, list[dict[str, Any]]], config: dict[str, Any]
) -> tuple[dict[str, Any], list[str]]:
    values = {}
    for record in normalized.get("financial_data", []):
        field_key = record.get("field_key")
        if field_key and record.get("value") not in (None, "", []):
            values[str(field_key)] = record["value"]
    configured, issues = resolve_configured_ocr_fields(normalized, config)
    values.update(configured)
    return values, issues


def _provider_fields(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    nested = payload.get("fields")
    return nested if isinstance(nested, dict) else payload


def _normalized_company_name(value: Any) -> str:
    """Normalize company names just enough for a provider identity check."""
    text = re.sub(r"[\s（）()\[\]【】]", "", str(value or ""))
    return text.replace("有限责任公司", "有限公司")


def _validated_qcc_payload(
    payload: Any,
    requested_name: str,
    role: str,
    issues: list[str],
) -> dict[str, Any]:
    """Reject a QCC response for a different company instead of filling it."""
    if not isinstance(payload, dict):
        return {}
    profile = payload.get("profile")
    returned_name = profile.get("name") if isinstance(profile, dict) else ""
    if returned_name and requested_name:
        requested = _normalized_company_name(requested_name)
        returned = _normalized_company_name(returned_name)
        if requested != returned:
            issues.append(
                f"企查查身份核验失败（{role}）：查询“{requested_name}”返回“{returned_name}”，相关字段已留空"
            )
            return {}
    return payload


def _asset_method_label(selected: Any) -> str:
    """Return the method label for the template's second result section."""
    text = str(selected or "")
    if "资产基础法" in text:
        return "资产基础法"
    if "市场法" in text:
        return "市场法"
    if "收益法" in text:
        return "收益法"
    return ""


def _filter_provider(
    payload: Any,
    allowed: set[str],
    provider_name: str,
    issues: list[str],
) -> dict[str, Any]:
    result = {}
    for field_key, value in _provider_fields(payload).items():
        if field_key not in allowed:
            issues.append(f"{provider_name} 返回越权字段，已丢弃：{field_key}")
        elif value not in (None, "", []):
            result[field_key] = value
    return result


def _paragraph_replacements(config: dict[str, Any], fields: dict[str, Any]) -> dict[tuple[str, int], str]:
    result = {}
    string_fields = defaultdict(
        lambda: "XXX",
        {key: str(value) for key, value in fields.items()},
    )
    for spec in config.get("paragraph_replacements", []):
        if "field_key" in spec:
            value = str(fields.get(spec["field_key"], ""))
        elif "template" in spec:
            blank_if = spec.get("blank_if_empty")
            if blank_if and not str(fields.get(blank_if, "") or "").strip():
                value = ""
            else:
                value = spec["template"].format_map(string_fields)
        else:
            value = str(spec.get("value", ""))
        # Do not leave empty manual-input brackets in an otherwise valid
        # report when a project intentionally omits an optional short name.
        value = value.replace("（简称：）", "").replace("（以下简称：）", "")
        result[(spec["part"], int(spec["paragraph_index"]))] = value
    return result


def _source_path(
    base: Path,
    config: dict[str, Any],
    overrides: dict[str, Path | None] | None,
    source_name: str,
) -> Path | None:
    if overrides is not None and source_name in overrides:
        override = overrides[source_name]
        return Path(override).resolve() if override is not None else None
    configured = config.get("sources", {}).get(source_name)
    return _path(base, configured) if configured else None


def _company_profile_table(profile: dict[str, Any], fallback_name: Any = "", fallback_capital: Any = "") -> list[list[str]]:
    """Render the backend-provided company profile into the template's 2-cell table."""
    profile = profile if isinstance(profile, dict) else {}
    present = lambda value: str(value) if value not in (None, "") else "XXX"
    credit_code = present(profile.get("credit_code"))
    name = present(profile.get("name") or fallback_name)
    capital = present(profile.get("registered_capital") or fallback_capital)
    return [
        [f"统一社会信用代码：{credit_code}", f"企业名称：{name}"],
        [f"类型：{present(profile.get('company_type'))}", f"法定代表人：{present(profile.get('legal_representative'))}"],
        [f"注册资本：{capital}", f"成立日期：{present(profile.get('establish_date'))}"],
        [f"营业期限自：{present(profile.get('term_start'))}", f"营业期限至：{present(profile.get('term_end'))}"],
        [f"登记机关：{present(profile.get('registration_authority'))}", f"核准日期：{present(profile.get('approval_date'))}"],
        [f"登记状态：{present(profile.get('status'))}"],
        [f"注册地址：{present(profile.get('address'))}"],
        [f"许可项目：{present(profile.get('business_scope'))}"],
    ]


def _amount_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return str(value).strip()


def _configured_cross_source_table(
    base: Path,
    config: dict[str, Any],
    rows: list[dict[str, Any]],
    ocr_overrides: dict[str, str] | None = None,
    source_overrides: dict[str, Path | None] | None = None,
) -> list[list[str]]:
    matrix = [["项目", "账面金额（元）", "数量", "现状、特点"]]
    for row in rows:
        source_name = row["source"]
        source_path = _source_path(
            base,
            config,
            source_overrides,
            source_name,
        )
        value = (ocr_overrides or {}).get(str(row.get("ocr_field_key")))
        if value in (None, ""):
            values, _ = try_read_cells(source_path, [row["locator"]])
            value = values.get(row["locator"], "XXX")
        matrix.append([str(row["label"]), _amount_text(value), str(row.get("quantity", "")), str(row.get("condition", ""))])
    return matrix


def _apply_ocr_overrides_to_table(
    matrix: list[list[str]], spec: dict[str, Any], overrides: dict[str, str]
) -> list[list[str]]:
    """Replace configured table amounts with semantically matched PDF OCR values."""
    by_label = {
        str(row.get("label", "")).strip(): str(row.get("ocr_field_key"))
        for row in spec.get("rows", [])
        if row.get("ocr_field_key")
    }
    for row in matrix:
        if not row:
            continue
        key = by_label.get(str(row[0]).strip())
        if key and overrides.get(key) not in (None, "") and len(row) > 1:
            row[1] = str(overrides[key])
    return matrix


def _qcc_table_rows(payload: dict[str, Any], key: str, width: int) -> list[list[str]]:
    rows = payload.get(key, []) if isinstance(payload, dict) else []
    matrix = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        if key == "trademark_rows":
            # The reviewed report uses the second column as an optional image
            # cell and the third column as the textual trademark name.  Do not
            # put an image URL (or the fallback word “图样”) into the name
            # column.  A graphical mark is represented explicitly as “图形”.
            name = row.get("name", "") or ("图形" if row.get("image") else "")
            values = [row.get("application_date", ""), "", name, row.get("registration_number", ""), row.get("class", ""), row.get("status", ""), row.get("announcement_date", "")]
        else:
            values = [row.get(k, "") for k in ("index", "name", "registration_number", "first_publication_date", "approval_date")]
        matrix.append([str(value or "") for value in values[:width]])
    return matrix


def _money_yuan(value: Any, *, plain_number_unit: str = "万") -> str:
    """Normalize QCC/reference capital values to the template's yuan unit."""
    if value in (None, ""):
        return ""
    text = str(value).strip().replace(",", "").replace("人民币", "")
    unit = 1
    if "万" in text or plain_number_unit == "万":
        unit = 10000
    text = text.replace("万元", "").replace("万", "").replace("元", "").strip()
    try:
        from decimal import Decimal

        return f"{Decimal(text) * unit:,.2f}"
    except Exception:
        return str(value).strip()


def _equity_matrix_from_partners(rows: list[dict[str, Any]]) -> list[list[str]]:
    matrix = [["序号", "股东名称", "总出资（元）", "股权比例"]]
    total = 0.0
    for index, row in enumerate(rows, 1):
        name = str(row.get("name", ""))
        capital = _money_yuan(row.get("capital", ""))
        percent = str(row.get("percent", ""))
        if not (name or capital or percent):
            continue
        matrix.append([str(index), name, capital, percent])
        try:
            total += float(capital.replace(",", ""))
        except (ValueError, AttributeError):
            pass
    if len(matrix) > 1:
        # Keep one empty input row before the total row.  This is part of the
        # supplied report layout (and is retained in the reviewed reference),
        # so a one-shareholder result does not collapse the table vertically.
        if len(matrix) == 2:
            matrix.append(["", "", "", ""])
        matrix.append(["合计", "合计", f"{total:,.2f}" if total else "", "100%"])
    return matrix


def _ocr_ownership_matrix(
    normalized: dict[str, list[dict[str, Any]]], spec: dict[str, Any]
) -> tuple[list[list[str]], list[str]]:
    """Build the founding-ownership table from configured PDF-OCR cells.

    Audit reports often split a shareholder name over several OCR rows and
    may place the amount in a neighbouring row.  The project configuration
    therefore declares the exact cells to join instead of relying on a
    positional guess.  This keeps the rule reusable for another report
    layout while preserving the PDF/XLSX source boundary.
    """
    matched_cells = find_ocr_table(
        normalized,
        table_id=str(spec.get("table_id", "")),
        table_markers=list(spec.get("table_markers", [])),
        page_markers=list(spec.get("page_markers", [])),
    )
    cells = {
        (int(cell.get("row", 0)), int(cell.get("column", 0))): str(cell.get("text") or "").strip()
        for cell in matched_cells
    }

    def cell(ref: Any) -> str:
        if not isinstance(ref, (list, tuple)) or len(ref) != 2:
            return ""
        try:
            return cells.get((int(ref[0]), int(ref[1])), "")
        except (TypeError, ValueError):
            return ""

    matrix = [list(spec.get("header", ["序号", "股东名称", "总出资（元）", "股权比例"]))]
    if not spec.get("rows"):
        return _auto_ocr_ownership_matrix(matched_cells, matrix)
    names: list[str] = []
    total = 0.0
    for index, row in enumerate(spec.get("rows", []), 1):
        name = "".join(cell(ref) for ref in row.get("name_cells", []))
        capital = cell(row.get("capital_cell"))
        percent = str(row.get("percent", ""))
        if not (name or capital or percent):
            continue
        names.append(name)
        matrix.append([str(index), name, capital, percent])
        try:
            total += float(capital.replace(",", ""))
        except (TypeError, ValueError):
            pass
    if len(matrix) > 1 and spec.get("include_total", True):
        if len(matrix) == 2 and spec.get("preserve_blank_row", True):
            matrix.append(["", "", "", ""])
        matrix.append(["合计", "合计", f"{total:,.2f}" if total else "", "100%"])
    return matrix, names


def _auto_ocr_ownership_matrix(
    cells: list[dict[str, Any]], matrix: list[list[str]]
) -> tuple[list[list[str]], list[str]]:
    """Infer shareholder rows when a PDF changes page/table coordinates."""
    from demo.domain.financial_matching import parse_number

    by_row: dict[int, list[dict[str, Any]]] = {}
    for cell in cells:
        by_row.setdefault(int(cell.get("row", 0)), []).append(cell)
    groups: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    for row_number in sorted(by_row):
        row = sorted(by_row[row_number], key=lambda item: int(item.get("column", 0)))
        first = "".join(str(cell.get("text") or "").strip() for cell in row if int(cell.get("column", 0)) == 1)
        if not current and (not first or first in {"项目", "出资方", "合计"}):
            continue
        if current or first:
            current.extend(row)
        if current and ("有限公司" in first or "公司" in first or "企业" in first):
            groups.append(current)
            current = []
    result_names: list[str] = []
    total = 0.0
    parsed: list[tuple[str, str]] = []
    for group in groups:
        name = "".join(
            str(cell.get("text") or "").strip()
            for cell in sorted(group, key=lambda item: (int(item.get("row", 0)), int(item.get("column", 0))))
            if int(cell.get("column", 0)) == 1 and parse_number(cell.get("text")) is None
        )
        amounts = [
            parse_number(cell.get("text"))
            for cell in group
            if parse_number(cell.get("text")) is not None
        ]
        if not name or not amounts:
            continue
        amount = float(amounts[0])
        parsed.append((name, f"{amount:,.2f}"))
        total += amount
    for index, (name, capital) in enumerate(parsed, 1):
        percent = f"{float(capital.replace(',', '')) / total:.0%}" if total else ""
        result_names.append(name)
        matrix.append([str(index), name, capital, percent])
    if parsed:
        matrix.append(["合计", "合计", f"{total:,.2f}", "100%"])
    return matrix, result_names


def _equity_matrix_from_reference(base: Path, config: dict[str, Any], table_index: int) -> list[list[str]]:
    source_path = _path(base, config["sources"]["reference_report"])
    source = read_table_matrix(source_path, int(table_index))
    matrix = [["序号", "股东名称", "总出资（元）", "股权比例"]]
    for row in source[1:]:
        values = [str(value or "").strip() for value in row]
        if len(values) < 4 or not any(values):
            continue
        matrix.append([values[0], values[1], _money_yuan(values[2]), values[3]])
    return matrix


def _profile_matrix_from_reference(base: Path, config: dict[str, Any], table_index: int) -> list[list[str]]:
    source_path = _path(base, config["sources"]["reference_report"])
    source = read_table_matrix(source_path, int(table_index))
    rows = [[str(value or "").strip() for value in row] for row in source if row]
    values: dict[str, str] = {}
    for row in rows:
        if not row:
            continue
        label = row[0]
        if label == "统一社会信用代码":
            values["credit_code"] = row[1] if len(row) > 1 else ""
            values["name"] = row[3] if len(row) > 3 else ""
        elif label in {"类型", "注册资本", "成立日期", "住所", "登记状态", "营业期限自", "营业期限至", "经营范围"}:
            values[label] = row[1] if len(row) > 1 else ""
            if label == "类型" and len(row) > 3:
                values["法定代表人"] = row[3]
            elif label == "注册资本" and len(row) > 3:
                values["成立日期"] = row[3]
            elif label == "营业期限自" and len(row) > 3:
                values["营业期限至"] = row[3]

    return [
        [f"统一社会信用代码：{values.get('credit_code', '')}", f"企业名称：{values.get('name', '')}"],
        [f"类型：{values.get('类型', '')}", f"法定代表人：{values.get('法定代表人', '')}"],
        [f"注册资本：{values.get('注册资本', '')}", f"成立日期：{values.get('成立日期', '')}"],
        [f"营业期限自：{values.get('营业期限自', '')}", f"营业期限至：{values.get('营业期限至', '')}"],
        ["登记机关：", "核准日期："],
        [f"登记状态：{values.get('登记状态', '')}"],
        [f"注册地址：{values.get('住所', '')}"],
        [f"许可项目：{values.get('经营范围', '')}"],
    ]


def run_pipeline(
    *,
    project_config: Path,
    pdf_path: Path | None,
    output_dir: Path,
    ocr_adapter: Any,
    llm_adapter: Any = None,
    qichacha_adapter: Any = None,
    node_inputs: dict[str, Any] | None = None,
    ocr_field_resolver: OcrFieldResolver | None = None,
    template_path: Path | None = None,
    template_page_reader: Any = None,
    report_date: str | None = None,
    manual_inputs_override: dict[str, Any] | None = None,
    ocr_workbook_path: Path | None = None,
    source_overrides: dict[str, Path | None] | None = None,
    review_adapters: dict[str, Any] | None = None,
    workflow_path: Path | None = None,
) -> PipelineResult:
    config_path = project_config.resolve()
    config = json.loads(config_path.read_text(encoding="utf-8"))
    workflow_file = (
        workflow_path.resolve()
        if workflow_path is not None
        else Path(__file__).with_name("workflow.yaml")
    )
    workflow_payload = json.loads(workflow_file.read_text(encoding="utf-8"))
    contract_result = validate_workflow_contract(workflow_payload, schemas)
    if not contract_result["valid"]:
        raise ValueError(
            "工作流契约校验失败：" + "；".join(contract_result["issues"])
        )
    workflow_definition = schemas.WorkflowDefinition.model_validate(workflow_payload)
    workflow_nodes = {node.name: node for node in workflow_definition.nodes}
    data_manifest_path = Path(__file__).with_name("data_manifest.yaml")
    data_manifest = json.loads(data_manifest_path.read_text(encoding="utf-8"))
    trace_versions = {
        name: str(version)
        for name, version in data_manifest.get("rule_versions", {}).items()
    }
    trace_versions["data_manifest"] = str(data_manifest.get("version", ""))
    recorder = WorkflowTraceRecorder(
        workflow_version=workflow_definition.version,
        contract_version=workflow_definition.contract_version,
        versions=trace_versions,
    )

    def record_node(
        name: str,
        input_payload: dict[str, Any],
        output_payload: dict[str, Any],
        *,
        status: str = "completed",
        node_evidence: list[dict[str, Any]] | None = None,
        node_issues: list[str] | None = None,
    ) -> None:
        definition = workflow_nodes[name]
        recorder.record(
            node_name=name,
            input_model=getattr(schemas, definition.input_model),
            output_model=getattr(schemas, definition.output_model),
            input_payload=input_payload,
            output_payload=output_payload,
            status=status,
            evidence=node_evidence or [],
            issues=node_issues or [],
            human_checkpoint=definition.human_checkpoint,
        )

    base = config_path.parent
    template = template_path.resolve() if template_path else _path(base, config["template"])
    pdf = pdf_path.resolve() if pdf_path is not None else None
    output_dir = output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    mapping_path = _path(base, config["mapping"])
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    locations = validate_mapping(mapping)
    static_locations = mapping.get("static_locations", [])
    field_names = {
        item["field_key"]: item["field_name"]
        for item in [*locations, *static_locations]
        if item.get("field_key")
    }
    routes = load_yellow_routes(config["yellow_routes"])
    template_inventory = inventory_template(template)
    inventory_trace = [
        {
            "location_id": item["location_id"],
            "record_type": item["record_type"],
            "context": item["context"],
            "marker": item["marker"],
        }
        for item in template_inventory
    ]
    record_node(
        "inventory",
        {"template_path": str(template)},
        {"locations": inventory_trace},
        node_evidence=[
            {
                "source_kind": "word_template",
                "source_file": template.name,
                "source_locator": "全文占位符及黄色标注",
            }
        ],
    )
    yellow_location_ids = {
        item["location_id"]
        for item in template_inventory
        if item["record_type"] == "黄色标注内容块"
    }
    validate_yellow_routes(routes, expected_location_ids=yellow_location_ids)

    template_hash = _sha256(template)
    issues: list[str] = []
    with tempfile.TemporaryDirectory(prefix="appraisal-base-") as temporary:
        legacy = run_project(
            config_path,
            output_dir=Path(temporary),
            offline=True,
            report_date=report_date,
            manual_inputs_override=manual_inputs_override,
            source_overrides=source_overrides,
        )
        fields = json.loads((Path(temporary) / "normalized_fields.json").read_text(encoding="utf-8"))
        evidence = _legacy_evidence(legacy.audit_path)

    # The valuation object is a controlled user input even though the current
    # template does not mark every occurrence in yellow.  Validate it once and
    # reuse the exact standard wording everywhere in the report.
    if fields.get("valuation_subject_type") not in (None, ""):
        fields["valuation_subject_type"] = validate_valuation_subject_type(fields["valuation_subject_type"])
    fields["asset_approach_method_label"] = _asset_method_label(
        fields.get("selected_valuation_method")
    )

    # A project may have a fixed-size summary table already present in the
    # template but not represented as a yellow location.  Load it explicitly
    # from its configured source so the template's defaults are never reused.
    scope_table = config.get("asset_scope_summary_table")
    if isinstance(scope_table, dict):
        source_name = str(scope_table["source"])
        source_path = _source_path(
            base,
            config,
            source_overrides,
            source_name,
        )
        scope_rows, scope_issues = try_read_configured_table(
            source_path,
            scope_table,
        )
        if scope_rows is None:
            semantic_scope = fields.get(scope_table["field_key"])
            if isinstance(semantic_scope, dict) and isinstance(
                semantic_scope.get("rows"), list
            ):
                scope_rows = semantic_scope["rows"]
                scope_issues = []
            else:
                scope_rows = blank_configured_table(
                    scope_table,
                    placeholder="XXX",
                )
        issues.extend(
            f"{scope_table['field_key']}：{message}"
            for message in scope_issues
        )
        fields[scope_table["field_key"]] = {
            "caption": scope_table.get("caption", ""),
            "rows": scope_rows,
        }
        if not (
            evidence.get(scope_table["field_key"], {}).get("kind")
            == "semantic_excel"
            and not scope_issues
        ):
            evidence[scope_table["field_key"]] = {
                "kind": (
                    config.get("source_lineage", {})
                    .get(source_name, {})
                    .get("kind", source_name)
                    if source_path is not None and not scope_issues
                    else "missing"
                ),
                "file": source_path.name if source_path is not None else "",
                "locator": scope_table.get("source_locator", ""),
            }

    ocr_node_issues: list[str] = []
    if ocr_workbook_path is not None:
        try:
            normalized = normalized_from_ocr_workbook(ocr_workbook_path.resolve())
        except Exception as exc:
            cache_issue = f"OCR 缓存读取失败：{exc}"
            issues.append(cache_issue)
            ocr_node_issues.append(cache_issue)
            if pdf is not None and ocr_adapter is not None:
                pages, ocr_issues = ocr_adapter.extract(pdf)
                issues.extend(ocr_issues)
                ocr_node_issues.extend(ocr_issues)
                normalized = normalize_ocr_pages(pages)
            else:
                normalized = normalize_ocr_pages([])
    elif pdf is not None:
        pages, ocr_issues = (
            ocr_adapter.extract(pdf)
            if ocr_adapter is not None
            else ([], ["OCR 未配置"])
        )
        issues.extend(ocr_issues)
        ocr_node_issues.extend(ocr_issues)
        normalized = normalize_ocr_pages(pages)
    else:
        normalized = normalize_ocr_pages([])

    page_counts: dict[int, int] = {}
    for record in [*normalized.get("text_blocks", []), *normalized.get("table_cells", [])]:
        try:
            page_number = int(record.get("page_number", 0))
            page_count = int(record.get("page_count", 0))
        except (TypeError, ValueError):
            continue
        if page_number > 0:
            page_counts[page_number] = max(page_count, len(page_counts))
    ocr_page_summary = [
        {
            "page_number": page_number,
            "page_count": page_counts[page_number],
            "blocks": [],
            "tables": [],
        }
        for page_number in sorted(page_counts)
    ]
    record_node(
        "ocr_pdf",
        {"pdf_path": str(pdf) if pdf is not None else ""},
        {
            "document": {
                "source_file": pdf.name if pdf is not None else "",
                "pages": ocr_page_summary,
                "issues": ocr_node_issues,
            }
        },
        status=(
            "skipped"
            if pdf is None and ocr_workbook_path is None
            else "completed_with_issues"
            if ocr_node_issues
            else "completed"
        ),
        node_evidence=(
            [
                {
                    "source_kind": (
                        "ocr_cache"
                        if ocr_workbook_path
                        else "pdf_ocr"
                    ),
                    "source_file": (
                        ocr_workbook_path.name
                        if ocr_workbook_path
                        else pdf.name
                        if pdf is not None
                        else ""
                    ),
                    "source_locator": "页级 OCR 结构",
                    "text_block_count": len(
                        normalized.get("text_blocks", [])
                    ),
                    "table_cell_count": len(
                        normalized.get("table_cells", [])
                    ),
                }
            ]
            if pdf is not None or ocr_workbook_path is not None
            else []
        ),
        node_issues=ocr_node_issues,
    )

    # The founding-shareholder table is explicitly marked as a PDF audit
    # report lookup in the template.  Keep it separate from the QCC current
    # shareholder table and derive the sentence immediately above it from the
    # same configured OCR cells.
    historical_ownership_matrix: list[list[str]] | None = None
    founding_names: list[str] = []
    ownership_table_spec = config.get("historical_ownership_table")
    if isinstance(ownership_table_spec, dict):
        historical_ownership_matrix, founding_names = _ocr_ownership_matrix(normalized, ownership_table_spec)
        if founding_names:
            fields["founding_shareholder_1"] = founding_names[0]
            fields["founding_shareholder_2"] = founding_names[1] if len(founding_names) > 1 else ""
            evidence["founding_shareholder_1"] = {
                "kind": "pdf_ocr_xlsx",
                "file": (
                    ocr_workbook_path.name
                    if ocr_workbook_path
                    else pdf.name
                    if pdf is not None
                    else ""
                ),
                "locator": f"OCR_表格!{ownership_table_spec.get('table_id', '')}",
            }
            evidence["founding_shareholder_2"] = dict(evidence["founding_shareholder_1"])
    resolver = ocr_field_resolver or _default_ocr_field_resolver
    resolved_ocr, resolver_issues = resolver(normalized, config)
    issues.extend(resolver_issues)

    ocr_allowed = fields_for_route(routes, RouteKind.PDF_OCR_XLSX)
    ocr_values = _filter_provider(resolved_ocr, ocr_allowed, "PDF OCR/XLSX 解析器", issues)
    ocr_aux_values = resolve_ocr_aux_fields(normalized, config)
    if isinstance(scope_table, dict) and scope_table.get("field_key") in fields:
        scope_rows = fields[scope_table["field_key"]].get("rows", [])
        fields[scope_table["field_key"]]["rows"] = _apply_ocr_overrides_to_table(
            scope_rows, scope_table, ocr_aux_values
        )
        if any(row.get("ocr_field_key") in ocr_aux_values for row in scope_table.get("rows", [])):
            evidence[scope_table["field_key"]] = {
                "kind": "pdf_ocr_xlsx",
                "file": "OCR结构化结果.xlsx",
                "locator": scope_table.get("source_locator", "") + "；语义 OCR 覆盖",
            }
    ocr_fallback_fields = set(config.get("ocr_fallback_fields", []))
    ocr_prefer_material_fields = set(config.get("ocr_prefer_material_fields", []))
    for field_key in ocr_allowed:
        if field_key in ocr_prefer_material_fields and fields.get(field_key) not in (None, "", []):
            ocr_values[field_key] = fields[field_key]
        if field_key in ocr_fallback_fields and field_key not in ocr_values and fields.get(field_key) not in (None, "", []):
            ocr_values[field_key] = fields[field_key]
        if field_key in ocr_values:
            source = evidence.get(field_key, {})
            normalized["financial_data"].append(
                {
                    "field_key": field_key,
                    "field_name": field_key,
                    "period": "",
                    "value": ocr_values[field_key],
                    "unit": "",
                    "evidence_id": source.get("locator", "ocr:xlsx"),
                }
            )
    ocr_workbook = (
        export_ocr_workbook(
            output_dir / "OCR结构化结果.xlsx",
            normalized,
        )
        if pdf is not None or ocr_workbook_path is not None
        else None
    )
    record_node(
        "export_ocr_workbook",
        {
            "normalized_ocr": {
                "text_blocks": [
                    {"count": len(normalized.get("text_blocks", []))}
                ],
                "table_cells": [
                    {"count": len(normalized.get("table_cells", []))}
                ],
                "financial_data": [
                    {"count": len(normalized.get("financial_data", []))}
                ],
                "issues": [{"count": len(normalized.get("issues", []))}],
            },
            "output_path": str(ocr_workbook) if ocr_workbook else "",
        },
        {"workbook_path": str(ocr_workbook) if ocr_workbook else ""},
        status="completed" if ocr_workbook else "skipped",
        node_evidence=(
            [
                {
                    "source_kind": "generated_artifact",
                    "source_file": ocr_workbook.name,
                    "source_locator": "OCR_文本、OCR_表格、标准财务数据、识别问题",
                }
            ]
            if ocr_workbook
            else []
        ),
    )

    configured_source_paths = {}
    for name in config.get("sources", {}):
        path = _source_path(base, config, source_overrides, name)
        if path is not None:
            configured_source_paths[name] = str(path)
    base_candidates = trace_candidates(fields, evidence, field_names)
    record_node(
        "extract_sources",
        {"sources": configured_source_paths},
        {
            "candidates": base_candidates,
            "issues": list(legacy.issues),
        },
        status="completed_with_issues" if legacy.issues else "completed",
        node_evidence=[
            {
                "source_kind": "source_materials",
                "source_file": Path(path).name,
                "source_locator": name,
            }
            for name, path in configured_source_paths.items()
        ],
        node_issues=list(legacy.issues),
    )

    manual_path = _path(base, config["manual_inputs"])
    configured_inputs = (
        {}
        if manual_inputs_override is not None
        else json.loads(manual_path.read_text(encoding="utf-8"))
        if manual_path.exists()
        else {}
    )
    if manual_inputs_override:
        configured_inputs.update({key: value for key, value in manual_inputs_override.items() if value not in (None, "")})
    node_allowed = fields_for_route(routes, RouteKind.NODE_INPUT)
    node_values = {
        field_key: value
        for field_key, value in configured_inputs.items()
        if field_key in node_allowed and value not in (None, "", [])
    }
    node_values.update(_filter_provider(node_inputs or {}, node_allowed, "节点输入", issues))
    if node_values.get("selected_valuation_method") not in (None, ""):
        node_values["selected_valuation_method"] = normalize_valuation_methods(
            node_values["selected_valuation_method"]
        )

    qcc_allowed = fields_for_route(routes, RouteKind.QICHACHA_API)
    qcc_values: dict[str, Any] = {}
    qcc_profiles: dict[str, dict[str, Any]] = {}
    qcc_payloads: dict[str, dict[str, Any]] = {}
    software_no_result = False
    if qichacha_adapter is not None:
        commissioning_name = str(fields.get("commissioning_party_name", ""))
        target_name = str(fields.get("target_company_name", ""))
        if commissioning_name:
            payload, provider_issues = qichacha_adapter.fetch(commissioning_name)
            issues.extend(provider_issues)
            software_no_result = software_no_result or any("接口 233 返回 201" in issue for issue in provider_issues)
            payload = _validated_qcc_payload(payload, commissioning_name, "委托人", issues)
            qcc_payloads["commissioning"] = payload
            qcc_profiles["commissioning"] = qcc_payloads["commissioning"].get("profile", {})
            qcc_values.update(_filter_provider(payload, {"commissioning_party_profile"}, "企查查 API（委托人）", issues))
        if target_name:
            payload, provider_issues = qichacha_adapter.fetch(target_name)
            issues.extend(provider_issues)
            software_no_result = software_no_result or any("接口 233 返回 201" in issue for issue in provider_issues)
            payload = _validated_qcc_payload(payload, target_name, "被评估单位", issues)
            qcc_payloads["target"] = payload
            qcc_profiles["target"] = qcc_payloads["target"].get("profile", {})
            qcc_values.update(_filter_provider(payload, qcc_allowed - {"commissioning_party_profile"}, "企查查 API（被评估单位）", issues))
    target_profile = qcc_profiles.get("target", {})
    if (
        fields.get("registered_capital") in (None, "", [])
        and target_profile.get("registered_capital") not in (None, "")
    ):
        fields["registered_capital"] = target_profile["registered_capital"]
        evidence["registered_capital"] = {
            "kind": "qichacha_api",
            "file": "企查查 API（735）",
            "locator": "企业工商详情.注册资本",
        }

    # The detailed IP records belong in the dedicated tables.  Keep the
    # yellow paragraph as a short, stable cross-reference instead of dumping
    # dozens of patent/trademark rows into one body paragraph.
    target_ip = qcc_payloads.get("target", {})
    patent_count = len(target_ip.get("patent_rows", [])) if isinstance(target_ip, dict) else 0
    trademark_count = len(target_ip.get("trademark_rows", [])) if isinstance(target_ip, dict) else 0
    if patent_count or trademark_count:
        details = []
        if patent_count:
            details.append(f"专利{patent_count}项")
        if trademark_count:
            details.append(f"商标{trademark_count}项")
        qcc_values["unrecorded_intangibles"] = "已查询到" + "、".join(details) + "，详细信息见下表。"
        fields["trademark_summary"] = "商标明细见下表。" if trademark_count else fields.get("trademark_summary", "")
        evidence["trademark_summary"] = {
            "kind": "qichacha_api",
            "file": "企查查 API（231）",
            "locator": "全国商标查询",
        }
    software_count = len(target_ip.get("software_rows", [])) if isinstance(target_ip, dict) else 0
    software_query_ok = bool(target_ip.get("software_query_ok")) if isinstance(target_ip, dict) else False
    if not software_count and (software_no_result or software_query_ok) and "software_copyrights" in qcc_allowed:
        qcc_values["software_copyrights"] = "未查询到软件著作权登记记录。"
        evidence["software_copyrights"] = {
            "kind": "qichacha_api",
            "file": "企查查 API（233）",
            "locator": "软件著作权查询（有效请求但无结果）",
        }

    llm_allowed = fields_for_route(routes, RouteKind.BAILIAN_GLM)
    selected_modules = normalize_narrative_modules(fields.get("narrative_modules"))
    # The company profile is always generated; the six report modules follow
    # the user's checkbox selection from the front end.
    llm_allowed = select_narrative_fields(llm_allowed, selected_modules)
    llm_values: dict[str, Any] = {}
    llm_source_kind = "bailian_glm"
    if llm_adapter is not None:
        structured_evidence = []
        for field_key, value in fields.items():
            source = evidence.get(field_key, {})
            if (
                value in (None, "", [], {})
                or source.get("kind") in {"missing", "bailian_glm"}
            ):
                continue
            if isinstance(value, (dict, list)):
                rendered = json.dumps(value, ensure_ascii=False, default=str)
            else:
                rendered = str(value)
            structured_evidence.append(
                {
                    "evidence_id": f"field:{field_key}",
                    "text": (
                        f"{field_names.get(field_key, field_key)}："
                        f"{rendered[:4000]}"
                    ),
                }
            )
        reference_report = _source_path(
            base,
            config,
            source_overrides,
            "reference_report",
        )
        if (
            reference_report is not None
            and reference_report.suffix.lower() == ".docx"
        ):
            try:
                structured_evidence.extend(
                    read_narrative_evidence(
                        reference_report,
                        source_name="reference_report",
                    )
                )
            except (OSError, ValueError, KeyError) as exc:
                issues.append(f"参考 Word 叙述证据读取失败：{exc}")
        if target_profile:
            structured_evidence.append(
                {
                    "evidence_id": "api:qichacha:target_profile",
                    "text": "被评估单位工商信息："
                    + json.dumps(target_profile, ensure_ascii=False, default=str),
                }
            )
        llm_evidence = {
            "selected_modules": selected_modules,
            "evidence": [
                {"evidence_id": item["evidence_id"], "text": item["text"]}
                for item in [*normalized["text_blocks"], *normalized["table_cells"]]
            ] + structured_evidence,
        }
        payload, provider_issues = llm_adapter.generate(llm_evidence)
        issues.extend(provider_issues)
        llm_values = _filter_provider(payload, llm_allowed, "百炼 GLM", issues)
    # Narrative fields are the only fields allowed to use a project-level local
    # fallback.  This keeps financial and legal facts fail-closed while allowing
    # a report to be generated when an external LLM returns an unusable or
    # partial payload.  Merge per field rather than only when the whole payload
    # is empty: GLM may return ``main_products`` while omitting one of the other
    # numbered overview slots, and a blank numbered slot is a formatting defect.
    if llm_adapter is not None and source_overrides is None:
        fallback = config.get("llm_fallback_fields", {})
        if isinstance(fallback, dict):
            fallback_values = {
                key: str(value)
                for key, value in fallback.items()
                if key in llm_allowed and value not in (None, "", [])
            }
            missing_fallback = {
                key: value for key, value in fallback_values.items()
                if key not in llm_values or llm_values.get(key) in (None, "", [])
            }
            if missing_fallback:
                llm_values.update(missing_fallback)
                llm_source_kind = "bailian_glm_evidence_fallback"
                issues.append("百炼 GLM 未返回全部授权叙述字段，缺失字段已使用项目配置的 PDF 证据化回填。")

    providers = {
        RouteKind.PDF_OCR_XLSX: (ocr_values, "pdf_ocr_xlsx"),
        RouteKind.QICHACHA_API: (qcc_values, "qichacha_api"),
        RouteKind.BAILIAN_GLM: (llm_values, llm_source_kind),
        RouteKind.NODE_INPUT: (node_values, "node_input"),
    }
    for route in routes:
        provider_values, source_kind = providers[route.route_kind]
        value = provider_values.get(route.field_key, "")
        fields[route.field_key] = value
        prior_source = evidence.get(route.field_key, {})
        preserve_material_locator = route.route_kind == RouteKind.PDF_OCR_XLSX and value
        evidence_fallback = source_kind == "bailian_glm_evidence_fallback" and value
        pdf_name = pdf.name if pdf is not None else ""
        if preserve_material_locator:
            evidence_file = prior_source.get("file", "") or pdf_name
        elif evidence_fallback:
            evidence_file = pdf_name
        elif source_kind == "qichacha_api" and value:
            evidence_file = "企查查 API（735/231/514/233）"
        else:
            evidence_file = ""
        evidence[route.field_key] = {
            "kind": source_kind if value not in (None, "", []) else "missing",
            "file": evidence_file,
            "locator": (
                prior_source.get("locator", "") or route.location_id
                if preserve_material_locator
                else "PDF OCR 证据化叙述回填"
                if evidence_fallback
                else route.location_id if value not in (None, "", []) else ""
            ),
        }
        if value in (None, "", []):
            issues.append(f"{route.field_key}：指定来源无可用值，已留空")

    # The template has a heading field and a separate body placeholder for the
    # same company-profile narrative.  Keep them synchronized so replacing the
    # yellow heading never leaves the original placeholder punctuation behind.
    if fields.get("company_profile_section") and not fields.get("company_profile_text"):
        fields["company_profile_text"] = str(fields["company_profile_section"]).rstrip("。；; ")
        evidence["company_profile_text"] = {
            "kind": "bailian_glm_profile_alias",
            "file": pdf.name if pdf is not None else "",
            "locator": "company_profile_section",
        }

    # The six overview slots are numbered sub-items in the template.  LLM
    # responses are allowed to include or omit the marker, but the generated
    # report must contain exactly one marker at each slot.
    for field_key, prefix in config.get("narrative_prefixes", {}).items():
        value = str(fields.get(field_key, "") or "").strip()
        if not value:
            continue
        value = re.sub(rf"^(?:{re.escape(prefix)})+", "", value).strip()
        fields[field_key] = prefix + value

    non_narrative_values = {
        field_key: value
        for field_key, value in fields.items()
        if field_key not in llm_allowed
    }
    non_narrative_candidates = trace_candidates(
        non_narrative_values,
        evidence,
        field_names,
    )
    non_narrative_resolved = trace_resolved_fields(
        non_narrative_values,
        evidence,
        field_names,
    )
    record_node(
        "resolve_fields",
        {
            "candidates": non_narrative_candidates,
            "mappings": trace_mappings(locations),
        },
        {"fields": non_narrative_resolved},
        status="completed_with_issues" if issues else "completed",
        node_evidence=[
            trace_evidence(evidence_item)
            for evidence_item in evidence.values()
            if isinstance(evidence_item, dict)
        ],
        node_issues=list(issues),
    )
    record_node(
        "select_narrative_modules",
        {"selected_modules": selected_modules},
        {"selected_modules": selected_modules},
        node_evidence=[
            {
                "source_kind": "node_input",
                "source_file": "人工基础信息",
                "source_locator": "主体概况模块",
            }
        ],
    )
    narrative_resolved = trace_resolved_fields(
        fields,
        evidence,
        field_names,
        keys=llm_allowed,
    )
    narrative_prompt_version = getattr(
        llm_adapter,
        "prompt_version",
        "yellow_narratives.v1",
    )
    recorder.versions["narrative_model"] = str(
        getattr(llm_adapter, "model", "") or ""
    )
    recorder.versions["narrative_prompt"] = str(narrative_prompt_version)
    record_node(
        "generate_narrative",
        {"fields": non_narrative_resolved},
        {
            "fields": narrative_resolved,
            "prompt_version": str(narrative_prompt_version),
        },
        status=(
            "skipped"
            if llm_adapter is None
            else "completed_with_issues"
            if any("LLM" in issue or "GLM" in issue for issue in issues)
            else "completed"
        ),
        node_evidence=[
            trace_evidence(evidence.get(field_key))
            for field_key in llm_allowed
            if fields.get(field_key) not in (None, "", [], {})
        ],
        node_issues=[
            issue
            for issue in issues
            if "LLM" in issue or "GLM" in issue
        ],
    )

    required_monetary_fields = list(config.get("required_monetary_fields", []))
    monetary_gate = require_financial_fields(fields, required_monetary_fields)
    monetary_policy = apply_missing_field_policy(
        fields,
        evidence,
        required_monetary_fields,
        "金额及财务结果字段",
    )
    fields = monetary_policy["fields"]
    evidence = monetary_policy["evidence"]
    financial_issues = list(monetary_policy["issues"])
    if monetary_gate["conflicts"]:
        financial_issues.append(
            "高优先级：金额及财务结果存在冲突，已保留待人工复核："
            + json.dumps(
                monetary_gate["conflicts"],
                ensure_ascii=False,
                default=str,
            )
        )
    issues.extend(financial_issues)
    financial_validation = {
        "valid": monetary_policy["valid"] and not monetary_gate["conflicts"],
        "missing_fields": monetary_policy["missing_fields"],
        "conflicts": monetary_gate["conflicts"],
    }

    replacements = build_replacements(locations, fields)
    table_replacements = {}
    table_column_ratios: dict[int, list[float]] = {}
    table_specs = list(config.get("financial_tables", []))
    scope_table = config.get("asset_scope_summary_table")
    if isinstance(scope_table, dict):
        table_specs.insert(0, scope_table)
    for spec in table_specs:
        value = fields.get(spec["field_key"])
        if isinstance(value, dict) and isinstance(value.get("rows"), list):
            table_replacements[int(spec["target_table_index"])] = value["rows"]
        else:
            table_replacements[int(spec["target_table_index"])] = blank_configured_table(
                spec,
                placeholder="XXX",
            )
        if str(spec.get("field_key", "")).startswith("historical_"):
            matrix = table_replacements[int(spec["target_table_index"])]
            column_count = len(matrix[0]) if matrix and matrix[0] else 0
            if column_count > 1:
                value_ratio = 0.68 / (column_count - 1)
                table_column_ratios[int(spec["target_table_index"])] = [
                    0.32,
                    *([value_ratio] * (column_count - 1)),
                ]
    if historical_ownership_matrix and len(historical_ownership_matrix) > 1:
        table_replacements[int(ownership_table_spec["target_table_index"])] = historical_ownership_matrix
    elif isinstance(ownership_table_spec, dict):
        table_replacements[int(ownership_table_spec["target_table_index"])] = [
            list(
                ownership_table_spec.get(
                    "header",
                    ["序号", "股东名称", "总出资（元）", "股权比例"],
                )
            ),
            ["XXX", "XXX", "XXX", "XXX"],
        ]
    # These tables are present in the communication template but were not
    # represented by yellow paragraphs.  Fill them explicitly so template
    # defaults (such as the listed-company capital) cannot leak into output.
    table_replacements[0] = _company_profile_table(
        qcc_profiles.get("commissioning", {}),
        fields.get("commissioning_party_name", ""),
    )
    table_replacements[1] = _company_profile_table(
        qcc_profiles.get("target", {}),
        fields.get("target_company_name", ""),
    )
    partner_rows = qcc_payloads.get("target", {}).get("partner_rows", [])
    if partner_rows:
        table_replacements[3] = _equity_matrix_from_partners(partner_rows)
    else:
        table_replacements[3] = [
            ["序号", "股东名称", "总出资（元）", "股权比例"],
            ["XXX", "XXX", "XXX", "XXX"],
        ]
    long_term_table = config.get("long_term_assets_table")
    if isinstance(long_term_table, dict):
        semantic_long_term = fields.get("long_term_assets_table")
        if isinstance(semantic_long_term, dict) and isinstance(
            semantic_long_term.get("rows"), list
        ):
            long_term_rows = _apply_ocr_overrides_to_table(
                semantic_long_term["rows"],
                long_term_table,
                ocr_aux_values,
            )
        else:
            long_term_rows = _configured_cross_source_table(
                base,
                config,
                long_term_table.get("rows", []),
                ocr_aux_values,
                source_overrides,
            )
        table_replacements[int(long_term_table["target_table_index"])] = long_term_rows
    trademark_rows = _qcc_table_rows(qcc_payloads.get("target", {}), "trademark_rows", 7)
    software_rows = _qcc_table_rows(qcc_payloads.get("target", {}), "software_rows", 5)
    if trademark_rows:
        table_replacements[8] = [["申请日期", "商标", "商标名称", "注册号", "国际分类", "商标状态", "注册公告日期"], *trademark_rows]
    else:
        table_replacements[8] = [
            ["申请日期", "商标", "商标名称", "注册号", "国际分类", "商标状态", "注册公告日期"],
            ["XXX", "XXX", "XXX", "XXX", "XXX", "XXX", "XXX"],
        ]
    if "software_copyrights" in qcc_allowed:
        table_replacements[9] = [["序号", "软件名称", "登记号", "首次发表日期", "登记批准日期"], ["", "", "", "", ""]]
    if software_rows:
        table_replacements[9] = [["序号", "软件名称", "登记号", "首次发表日期", "登记批准日期"], *software_rows]
    elif software_no_result and "software_copyrights" in qcc_allowed:
        table_replacements[9] = [["序号", "软件名称", "登记号", "首次发表日期", "登记批准日期"], ["", "未查询到软件著作权登记记录。", "", "", ""]]
    elif "software_copyrights" in qcc_allowed:
        table_replacements[9] = [
            ["序号", "软件名称", "登记号", "首次发表日期", "登记批准日期"],
            ["XXX", "XXX", "XXX", "XXX", "XXX"],
        ]

    report = output_dir / "资产评估报告_待复核.docx"
    audit = output_dir / "字段审计清单.xlsx"
    fill_template(
        template,
        report,
        replacements,
        table_replacements=table_replacements,
        table_column_ratios=table_column_ratios,
        paragraph_replacements=_paragraph_replacements(config, fields),
        replacement_modes={route.location_id: route.replacement_mode for route in routes},
    )
    replace_image_markers(report)
    replace_report_number_year(report, fields.get("report_number_year"))
    unresolved_findings = highlight_unresolved_placeholders(report)
    generation_issues = issues_from_word_findings(
        unresolved_findings,
        [*locations, *static_locations],
        fields,
        evidence,
    )
    if _sha256(template) != template_hash:
        raise RuntimeError("模板被意外修改")
    record_node(
        "fill_word",
        {
            "template_path": str(template),
            "output_path": str(report),
            "fields": trace_resolved_fields(fields, evidence, field_names),
        },
        {
            "report_path": str(report),
            "replacement_count": len(replacements),
        },
        node_evidence=[
            {
                "source_kind": "word_template",
                "source_file": template.name,
                "source_locator": "映射位置与表格",
            }
        ],
        status=(
            "completed"
            if financial_validation["valid"]
            else "completed_with_issues"
        ),
        node_issues=financial_issues,
    )
    template_pages: dict[str, int | str] = {}
    generated_pages: dict[str, int | str] = {}
    if template_page_reader is not None:
        try:
            report_page_texts, report_page_issues = (
                template_page_reader.extract(report)
            )
            issues.extend(report_page_issues)
            generated_pages = map_location_pages(
                unresolved_findings,
                report_page_texts,
                document_paragraph_texts(report),
            )
        except Exception as exc:
            issues.append(f"生成报告页码解析失败：{exc}")
        try:
            template_page_texts, template_page_issues = (
                template_page_reader.extract(template)
            )
            issues.extend(template_page_issues)
            template_pages = map_location_pages(
                [*locations, *static_locations],
                template_page_texts,
                document_paragraph_texts(template),
            )
            template_issue_pages = map_location_pages(
                unresolved_findings,
                template_page_texts,
                document_paragraph_texts(template),
            )
        except Exception as exc:
            issues.append(f"模板页码解析失败：{exc}")
            template_issue_pages = {}
    else:
        template_issue_pages = {}
    generation_issues = apply_page_locations(
        generation_issues,
        generated_pages,
        template_issue_pages,
    )
    issue_workbook = export_generation_issues(
        output_dir / "生成问题清单.xlsx",
        generation_issues,
    )
    issue_json = write_json(
        output_dir / "生成问题清单.json",
        generation_issues,
    )
    for issue in generation_issues:
        page = issue.get("page_number") or "页码不可用"
        issues.append(
            f"Word第{page}页 {issue.get('location_description', '')}："
            f"{issue.get('problem', '')}"
        )
    export_audit(
        audit,
        [*locations, *static_locations],
        fields,
        evidence,
        template_pages=template_pages,
    )
    reviews: dict[str, dict[str, Any]] = {}
    review_output_paths: list[str] = []
    review_specs = {
        "format": {
            "node_name": "llm_format_review",
            "review_type": "format_review",
            "file_name": "格式审核.json",
            "prompt_version": "review_format.v1",
        },
        "data": {
            "node_name": "llm_data_validation",
            "review_type": "data_validation",
            "file_name": "数据校验.json",
            "prompt_version": "review_data.v1",
        },
        "semantic": {
            "node_name": "llm_semantic_review",
            "review_type": "semantic_review",
            "file_name": "语义审核.json",
            "prompt_version": "review_semantic.v1",
        },
    }
    review_inputs: dict[str, dict[str, Any]] = {}
    if review_adapters:
        review_inputs = {
            "format": build_format_review_evidence(template, report),
            "data": build_data_review_evidence(report, fields, audit, evidence),
            "semantic": build_semantic_review_evidence(report, fields),
        }
    for review_name, spec in review_specs.items():
        adapter = (review_adapters or {}).get(review_name)
        if adapter is None:
            skipped_review = {
                "review_type": spec["review_type"],
                "status": "skipped",
                "summary": "未启用对应 LLM 审核",
                "findings": [],
                "model": "",
                "prompt_version": spec["prompt_version"],
            }
            record_node(
                spec["node_name"],
                {
                    "review_type": spec["review_type"],
                    "report_path": str(report),
                    "evidence": {},
                },
                skipped_review,
                status="skipped",
                node_issues=["未启用对应 LLM 审核"],
            )
            continue
        try:
            review_result, review_issues = adapter.review(review_inputs[review_name])
            reviews[review_name] = review_result
            issues.extend(review_issues)
            for finding in review_result.get("findings", []):
                location = finding.get("location", "")
                problem = finding.get("problem", "")
                issues.append(f"LLM {review_name}审核：{location}：{problem}".strip("："))
            review_path = write_json(output_dir / spec["file_name"], review_result)
            review_output_paths.append(str(review_path))
            review_status = str(review_result.get("status", "completed"))
            if review_status not in {
                "completed",
                "completed_with_issues",
                "skipped",
                "failed",
            }:
                review_status = (
                    "completed_with_issues"
                    if review_result.get("findings")
                    else "completed"
                )
            recorder.versions[f"{review_name}_review_model"] = str(
                review_result.get("model", "")
            )
            recorder.versions[f"{review_name}_review_prompt"] = str(
                review_result.get("prompt_version", spec["prompt_version"])
            )
            record_node(
                spec["node_name"],
                {
                    "review_type": spec["review_type"],
                    "report_path": str(report),
                    "evidence": review_inputs[review_name],
                },
                review_result,
                status=review_status,
                node_evidence=[
                    {
                        "source_kind": "generated_report",
                        "source_file": report.name,
                        "source_locator": spec["review_type"],
                    }
                ],
                node_issues=review_issues,
            )
        except Exception as exc:
            failed_review = {
                "review_type": spec["review_type"],
                "status": "failed",
                "summary": f"审核执行失败：{exc}",
                "findings": [],
                "model": str(getattr(adapter, "model", "")),
                "prompt_version": str(
                    getattr(adapter, "prompt_version", spec["prompt_version"])
                ),
            }
            reviews[review_name] = failed_review
            issue = f"LLM {review_name}审核失败：{exc}"
            issues.append(issue)
            review_path = write_json(output_dir / spec["file_name"], failed_review)
            review_output_paths.append(str(review_path))
            record_node(
                spec["node_name"],
                {
                    "review_type": spec["review_type"],
                    "report_path": str(report),
                    "evidence": review_inputs.get(review_name, {}),
                },
                failed_review,
                status="failed",
                node_issues=[issue],
            )
    aggregate = aggregate_reviews(reviews)
    record_node(
        "review_aggregate",
        {"reviews": reviews},
        aggregate,
        status=(
            "completed_with_issues"
            if aggregate["status"] == "completed_with_issues"
            else "completed"
        ),
        node_issues=[
            issue for issue in issues if issue.startswith("LLM ") and "审核" in issue
        ],
    )
    if reviews:
        review_output_paths.append(
            str(write_json(output_dir / "审核汇总.json", aggregate))
        )
    if (
        not generation_issues
        and should_create_candidate_report(
            reviews,
            financial_fields_complete=financial_validation["valid"],
        )
    ):
        final_report = output_dir / "资产评估报告_最终候选.docx"
        shutil.copy2(report, final_report)
        review_output_paths.append(str(final_report))
    write_json(output_dir / "normalized_fields.json", fields)
    write_json(output_dir / "issues.json", issues)
    planned_manifest_path = output_dir / "run_manifest.json"
    record_node(
        "export_audit",
        {
            "report_path": str(report),
            "fields": trace_resolved_fields(fields, evidence, field_names),
        },
        {
            "audit_path": str(audit),
            "manifest_path": str(planned_manifest_path),
        },
        status="completed_with_issues" if issues else "completed",
        node_evidence=[
            {
                "source_kind": "generated_artifact",
                "source_file": audit.name,
                "source_locator": "填充结果",
            }
        ],
        node_issues=list(issues),
    )
    trace_path = recorder.export(output_dir / "workflow_trace.json")
    manifest = {
        "project_id": config["project_id"],
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "template": str(template),
        "template_sha256": template_hash,
        "pdf": str(pdf) if pdf is not None else "",
        "pdf_sha256": _sha256(pdf) if pdf is not None else "",
        "mapping_version": "1.0.0",
        "yellow_route_version": "yellow_routes.v1",
        "financial_rule_version": "financial_aliases.v1",
        "prompt_version": getattr(llm_adapter, "prompt_version", "yellow_narratives.v1"),
        "llm_models": {
            name: review.get("model", "") for name, review in reviews.items()
        },
        "ocr_cache_reused": bool(ocr_workbook_path),
        "ocr_cache_source": str(ocr_workbook_path) if ocr_workbook_path else "",
        "workflow_version": workflow_definition.version,
        "workflow_contract_version": workflow_definition.contract_version,
        "financial_validation": financial_validation,
        "generation_validation": {
            "valid": not generation_issues,
            "unresolved_count": len(generation_issues),
        },
        "outputs": [
            *([str(ocr_workbook)] if ocr_workbook else []),
            str(report),
            str(audit),
            str(issue_workbook),
            str(issue_json),
            str(trace_path),
            *review_output_paths,
        ],
        "reviews": {
            name: {
                "status": review.get("status", ""),
                "model": review.get("model", ""),
                "prompt_version": review.get("prompt_version", ""),
            }
            for name, review in reviews.items()
        },
    }
    manifest_path = write_json(planned_manifest_path, manifest)
    return PipelineResult(report, audit, ocr_workbook, manifest_path, issues)
